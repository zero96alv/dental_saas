from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User, Group
from core.mixins import TenantLoginRequiredMixin, tenant_login_required, TenantSuccessUrlMixin
from django.conf import settings
from django.db import models
from django.db.models import Count, Sum, Avg, Min, Max, F, Q
from django.db.models.functions import ExtractWeek, ExtractYear, TruncMonth
from django.http import JsonResponse, HttpResponse, Http404, FileResponse
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
import json
from django.views.generic import (
    ListView,
    DetailView,
    CreateView,
    UpdateView,
    DeleteView,
    TemplateView
)
from django.urls import reverse_lazy
from django.contrib.messages.views import SuccessMessageMixin
from django.db import transaction
from django.contrib import messages
from django.template.loader import render_to_string
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import mm
import string
import random
from datetime import timedelta
from openpyxl import Workbook
from datetime import datetime, timedelta
from django.shortcuts import redirect
from django.contrib.auth import logout
from django.contrib.auth.views import LogoutView as DjangoLogoutView
from django.http import HttpResponseRedirect
from django.urls import reverse
import logging

# Importar forms y models de manera controlada para evitar ciclos
from . import forms
from . import models

logger = logging.getLogger(__name__)

# === SAT Catalogs CRUD ===
class SatFormaPagoListView(TenantLoginRequiredMixin, ListView):
    model = models.SatFormaPago
    template_name = 'core/configuracion/sat_forma_pago_list.html'
    context_object_name = 'items'

class SatFormaPagoCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.SatFormaPago
    fields = ['codigo', 'descripcion', 'activo']
    template_name = 'core/configuracion/sat_catalog_form.html'
    success_url = reverse_lazy('core:sat_forma_pago_list')
    success_message = 'Forma de Pago creada con éxito.'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Nueva Forma de Pago (SAT)'
        return ctx

class SatFormaPagoUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.SatFormaPago
    fields = ['codigo', 'descripcion', 'activo']
    template_name = 'core/configuracion/sat_catalog_form.html'
    success_url = reverse_lazy('core:sat_forma_pago_list')
    success_message = 'Forma de Pago actualizada con éxito.'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Editar Forma de Pago (SAT)'
        return ctx

class SatFormaPagoDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = models.SatFormaPago
    template_name = 'core/configuracion/sat_catalog_confirm_delete.html'
    success_url = reverse_lazy('core:sat_forma_pago_list')
    success_message = 'Forma de Pago eliminada con éxito.'

class SatMetodoPagoListView(TenantLoginRequiredMixin, ListView):
    model = models.SatMetodoPago
    template_name = 'core/configuracion/sat_metodo_pago_list.html'
    context_object_name = 'items'

class SatMetodoPagoCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.SatMetodoPago
    fields = ['codigo', 'descripcion', 'activo']
    template_name = 'core/configuracion/sat_catalog_form.html'
    success_url = reverse_lazy('core:sat_metodo_pago_list')
    success_message = 'Método de Pago creado con éxito.'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Nuevo Método de Pago (SAT)'
        return ctx

class SatMetodoPagoUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.SatMetodoPago
    fields = ['codigo', 'descripcion', 'activo']
    template_name = 'core/configuracion/sat_catalog_form.html'
    success_url = reverse_lazy('core:sat_metodo_pago_list')
    success_message = 'Método de Pago actualizado con éxito.'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Editar Método de Pago (SAT)'
        return ctx

class SatMetodoPagoDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = models.SatMetodoPago
    template_name = 'core/configuracion/sat_catalog_confirm_delete.html'
    success_url = reverse_lazy('core:sat_metodo_pago_list')
    success_message = 'Método de Pago eliminado con éxito.'

class SatRegimenFiscalListView(TenantLoginRequiredMixin, ListView):
    model = models.SatRegimenFiscal
    template_name = 'core/configuracion/sat_regimen_fiscal_list.html'
    context_object_name = 'items'

class SatRegimenFiscalCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.SatRegimenFiscal
    fields = ['codigo', 'descripcion', 'persona_fisica', 'persona_moral', 'activo']
    template_name = 'core/configuracion/sat_catalog_form.html'
    success_url = reverse_lazy('core:sat_regimen_fiscal_list')
    success_message = 'Régimen Fiscal creado con éxito.'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Nuevo Régimen Fiscal (SAT)'
        return ctx

class SatRegimenFiscalUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.SatRegimenFiscal
    fields = ['codigo', 'descripcion', 'persona_fisica', 'persona_moral', 'activo']
    template_name = 'core/configuracion/sat_catalog_form.html'
    success_url = reverse_lazy('core:sat_regimen_fiscal_list')
    success_message = 'Régimen Fiscal actualizado con éxito.'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Editar Régimen Fiscal (SAT)'
        return ctx

class SatRegimenFiscalDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = models.SatRegimenFiscal
    template_name = 'core/configuracion/sat_catalog_confirm_delete.html'
    success_url = reverse_lazy('core:sat_regimen_fiscal_list')
    success_message = 'Régimen Fiscal eliminado con éxito.'

class SatUsoCFDIListView(TenantLoginRequiredMixin, ListView):
    model = models.SatUsoCFDI
    template_name = 'core/configuracion/sat_uso_cfdi_list.html'
    context_object_name = 'items'

class SatUsoCFDICreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.SatUsoCFDI
    fields = ['codigo', 'descripcion', 'persona_fisica', 'persona_moral', 'activo']
    template_name = 'core/configuracion/sat_catalog_form.html'
    success_url = reverse_lazy('core:sat_uso_cfdi_list')
    success_message = 'Uso CFDI creado con éxito.'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Nuevo Uso CFDI (SAT)'
        return ctx

class SatUsoCFDIUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.SatUsoCFDI
    fields = ['codigo', 'descripcion', 'persona_fisica', 'persona_moral', 'activo']
    template_name = 'core/configuracion/sat_catalog_form.html'
    success_url = reverse_lazy('core:sat_uso_cfdi_list')
    success_message = 'Uso CFDI actualizado con éxito.'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Editar Uso CFDI (SAT)'
        return ctx

class SatUsoCFDIDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = models.SatUsoCFDI
    template_name = 'core/configuracion/sat_catalog_confirm_delete.html'
    success_url = reverse_lazy('core:sat_uso_cfdi_list')
    success_message = 'Uso CFDI eliminado con éxito.'

# === DASHBOARD FINANCIERO INTEGRAL ===
class DashboardFinancieroView(TenantLoginRequiredMixin, TemplateView):
    template_name = 'core/dashboard_financiero.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener fechas para filtros
        hoy = timezone.now().date()
        inicio_mes = hoy.replace(day=1)
        inicio_ano = hoy.replace(month=1, day=1)
        
        # === MÉTRICAS PRINCIPALES ===
        context.update(self.get_metricas_principales(hoy, inicio_mes, inicio_ano))
        
        # === GRÁFICOS Y TENDENCIAS ===
        context.update(self.get_datos_graficos(hoy, inicio_mes))
        
        # === ALERTAS Y PENDIENTES ===
        context.update(self.get_alertas_pendientes())
        
        # === PAGOS RECIENTES ===
        context.update(self.get_pagos_recientes())
        
        # === REPORTES RÁPIDOS ===
        context.update(self.get_accesos_reportes())
        
        return context
    
    def get_metricas_principales(self, hoy, inicio_mes, inicio_ano):
        """Obtiene las métricas principales del dashboard"""
        try:
            # Ingresos del día
            ingresos_hoy = models.Pago.objects.filter(
                fecha_pago__date=hoy
            ).aggregate(total=Sum('monto'))['total'] or 0
            
            # Ingresos del mes
            ingresos_mes = models.Pago.objects.filter(
                fecha_pago__date__gte=inicio_mes
            ).aggregate(total=Sum('monto'))['total'] or 0
            
            # Ingresos del año
            ingresos_ano = models.Pago.objects.filter(
                fecha_pago__date__gte=inicio_ano
            ).aggregate(total=Sum('monto'))['total'] or 0
            
            # Saldos pendientes totales
            saldos_pendientes = models.Cliente.objects.aggregate(
                total=Sum('saldo_pendiente')
            )['total'] or 0
            
            # Pagos del día (cantidad)
            pagos_hoy_count = models.Pago.objects.filter(
                fecha_pago__date=hoy
            ).count()
            
            # Pacientes con saldo pendiente
            pacientes_pendientes = models.Cliente.objects.filter(
                saldo_pendiente__gt=0
            ).count()
            
            # Citas pendientes de pago
            citas_pendientes = models.Cita.objects.filter(
                estado='completada',
                pagada=False
            ).count()
            
            # Promedio de pago diario del mes
            dias_transcurridos = (hoy - inicio_mes).days + 1
            promedio_diario = ingresos_mes / dias_transcurridos if dias_transcurridos > 0 else 0
            
            # Comparación con mes anterior
            inicio_mes_anterior = (inicio_mes - timedelta(days=1)).replace(day=1)
            fin_mes_anterior = inicio_mes - timedelta(days=1)
            
            ingresos_mes_anterior = models.Pago.objects.filter(
                fecha_pago__date__gte=inicio_mes_anterior,
                fecha_pago__date__lte=fin_mes_anterior
            ).aggregate(total=Sum('monto'))['total'] or 0
            
            # Calcular porcentaje de crecimiento
            if ingresos_mes_anterior > 0:
                crecimiento_mes = ((ingresos_mes - ingresos_mes_anterior) / ingresos_mes_anterior) * 100
            else:
                crecimiento_mes = 100 if ingresos_mes > 0 else 0
            
            return {
                'ingresos_hoy': ingresos_hoy,
                'ingresos_mes': ingresos_mes,
                'ingresos_ano': ingresos_ano,
                'saldos_pendientes': saldos_pendientes,
                'pagos_hoy_count': pagos_hoy_count,
                'pacientes_pendientes': pacientes_pendientes,
                'citas_pendientes': citas_pendientes,
                'promedio_diario': promedio_diario,
                'crecimiento_mes': crecimiento_mes,
                'ingresos_mes_anterior': ingresos_mes_anterior,
            }
        except Exception as e:
            logger.error(f"Error calculando métricas financieras: {e}")
            return {}
    
    def get_datos_graficos(self, hoy, inicio_mes):
        """Obtiene datos para gráficos y tendencias"""
        try:
            # Ingresos por día (últimos 30 días)
            hace_30_dias = hoy - timedelta(days=30)
            ingresos_diarios = []
            labels_dias = []
            
            for i in range(30):
                fecha = hace_30_dias + timedelta(days=i)
                ingreso = models.Pago.objects.filter(
                    fecha_pago__date=fecha
                ).aggregate(total=Sum('monto'))['total'] or 0
                
                ingresos_diarios.append(float(ingreso))
                labels_dias.append(fecha.strftime('%d/%m'))
            
            # Métodos de pago más utilizados
            metodos_pago = models.Pago.objects.filter(
                fecha_pago__date__gte=inicio_mes
            ).values('metodo_pago').annotate(
                total=Sum('monto'),
                cantidad=Count('id')
            ).order_by('-total')[:5]
            
            # Servicios más rentables
            servicios_rentables = models.ServicioCita.objects.filter(
                cita__fecha_hora__date__gte=inicio_mes,
                cita__pagada=True
            ).values('servicio__nombre').annotate(
                total_ingresos=Sum('precio_final'),
                cantidad=Count('id')
            ).order_by('-total_ingresos')[:5]
            
            # Ingresos por dentista (mes actual)
            ingresos_dentistas = models.Pago.objects.filter(
                fecha_pago__date__gte=inicio_mes,
                cita__dentista__isnull=False
            ).values(
                'cita__dentista__first_name',
                'cita__dentista__last_name'
            ).annotate(
                total=Sum('monto')
            ).order_by('-total')[:10]
            
            return {
                'ingresos_diarios': ingresos_diarios,
                'labels_dias': labels_dias,
                'metodos_pago': list(metodos_pago),
                'servicios_rentables': list(servicios_rentables),
                'ingresos_dentistas': list(ingresos_dentistas),
            }
        except Exception as e:
            logger.error(f"Error obteniendo datos de gráficos: {e}")
            return {}
    
    def get_alertas_pendientes(self):
        """Obtiene alertas y pendientes críticos"""
        try:
            # Pacientes con saldo alto
            pacientes_saldo_alto = models.Cliente.objects.filter(
                saldo_pendiente__gt=5000  # Más de $5000
            ).order_by('-saldo_pendiente')[:5]
            
            # Citas sin pagar (más de 7 días)
            hace_una_semana = timezone.now() - timedelta(days=7)
            citas_vencidas = models.Cita.objects.filter(
                estado='completada',
                pagada=False,
                fecha_hora__lt=hace_una_semana
            ).select_related('paciente', 'dentista').order_by('-fecha_hora')[:5]
            
            # Pagos del día que requieren atención
            pagos_altos_hoy = models.Pago.objects.filter(
                fecha_pago__date=timezone.now().date(),
                monto__gt=2000
            ).select_related('cita__paciente').order_by('-monto')[:5]
            
            return {
                'pacientes_saldo_alto': pacientes_saldo_alto,
                'citas_vencidas': citas_vencidas,
                'pagos_altos_hoy': pagos_altos_hoy,
            }
        except Exception as e:
            logger.error(f"Error obteniendo alertas: {e}")
            return {}
    
    def get_pagos_recientes(self):
        """Obtiene los pagos más recientes"""
        try:
            pagos_recientes = models.Pago.objects.select_related(
                'cita__paciente',
                'cita__dentista'
            ).order_by('-fecha_pago')[:10]
            
            return {
                'pagos_recientes': pagos_recientes,
            }
        except Exception as e:
            logger.error(f"Error obteniendo pagos recientes: {e}")
            return {}
    
    def get_accesos_reportes(self):
        """Obtiene accesos rápidos a reportes"""
        return {
            'reportes_disponibles': [
                {
                    'nombre': 'Reporte de Ingresos',
                    'descripcion': 'Análisis detallado de ingresos por período',
                    'url': 'core:reporte_ingresos',
                    'icon': 'fas fa-chart-line'
                },
                {
                    'nombre': 'Saldos Pendientes',
                    'descripcion': 'Pacientes con pagos pendientes',
                    'url': 'core:reporte_saldos',
                    'icon': 'fas fa-exclamation-triangle'
                },
                {
                    'nombre': 'Reporte de Facturación',
                    'descripcion': 'Facturación y comprobantes fiscales',
                    'url': 'core:reporte_facturacion',
                    'icon': 'fas fa-file-invoice'
                },
                {
                    'nombre': 'Servicios más Vendidos',
                    'descripcion': 'Análisis de servicios por demanda',
                    'url': 'core:reporte_servicios_vendidos',
                    'icon': 'fas fa-trophy'
                }
            ]
        }

# Fallback simple para "Enviar a caja"
@tenant_login_required
@require_POST
def enviar_a_caja_api(request, cita_id):
    cita = get_object_or_404(models.Cita, id=cita_id)
    cita.estado = 'ATN'
    cita.save(update_fields=['estado'])
    return JsonResponse({'success': True})

class AjaxLoginRequiredMixin(TenantLoginRequiredMixin):
    """
    Mixin para vistas basadas en clases que devuelve un JsonResponse 403
    si la solicitud es AJAX y el usuario no está autenticado,
    en lugar de redirigir a la página de login.
    """
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated and request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'error': 'Autenticación requerida. Por favor, inicie sesión de nuevo.'}, status=403)
        return super().dispatch(request, *args, **kwargs)

class DashboardView(TenantLoginRequiredMixin, TemplateView):
    def get_template_names(self):
        user = self.request.user
        if hasattr(user, 'paciente_perfil'):
            return ['core/portal/dashboard.html']
        
        if user.groups.filter(name='Administrador').exists() or user.is_superuser:
            return ['core/dashboards/admin_dashboard.html']
        elif user.groups.filter(name='Dentista').exists():
            return ['core/dashboards/dentista_dashboard.html']
        elif user.groups.filter(name='Recepcionista').exists():
            return ['core/dashboards/recepcionista_dashboard.html']
        else:
            return ['core/dashboards/admin_dashboard.html']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        if hasattr(user, 'paciente_perfil'):
            paciente = user.paciente_perfil
            context['paciente'] = paciente
            context['proximas_citas'] = models.Cita.objects.filter(
                paciente=paciente, 
                fecha_hora__gte=timezone.now()
            ).order_by('fecha_hora')
            context['saldo_total_pendiente'] = paciente.saldo_global or 0
            return context

        periodo = self.request.GET.get('periodo', 'hoy')
        today = timezone.now().date()
        
        if periodo == 'semana':
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
            context['periodo_seleccionado'] = 'Esta Semana'
        elif periodo == 'mes':
            start_date = today.replace(day=1)
            next_month = start_date.replace(day=28) + timedelta(days=4)
            end_date = next_month - timedelta(days=next_month.day)
            context['periodo_seleccionado'] = 'Este Mes'
        elif periodo == 'historico':
            start_date = None
            end_date = today
            context['periodo_seleccionado'] = 'Histórico'
        else:
            start_date = today
            end_date = today
            context['periodo_seleccionado'] = 'Hoy'
            
        context['periodo_param'] = periodo
        
        citas_periodo = models.Cita.objects.all()
        pagos_periodo = models.Pago.objects.all()
        pacientes_periodo = models.Paciente.objects.all()
        
        if start_date:
            citas_periodo = citas_periodo.filter(fecha_hora__date__gte=start_date, fecha_hora__date__lte=end_date)
            pagos_periodo = pagos_periodo.filter(fecha_pago__date__gte=start_date, fecha_pago__date__lte=end_date)
            pacientes_periodo = pacientes_periodo.filter(creado_en__date__gte=start_date, creado_en__date__lte=end_date)

        context['citas_total_periodo'] = citas_periodo.count()
        context['citas_pendientes_periodo'] = citas_periodo.filter(estado__in=['PRO', 'CON']).count()
        
        if user.groups.filter(name='Administrador').exists() or user.is_superuser:
            context['ingresos_periodo'] = pagos_periodo.aggregate(total=Sum('monto'))['total'] or 0
            
            # USAR SALDO_GLOBAL DE PACIENTES
            context['total_saldos_pendientes'] = models.Paciente.objects.filter(
                saldo_global__gt=0
            ).aggregate(total=Sum('saldo_global'))['total'] or 0
            
            from datetime import date
            alert_date = date.today()
            proximos_30_dias = alert_date + timedelta(days=30)
            context['stocks_bajos'] = models.Insumo.objects.filter(stock__lte=F('stock_minimo'))
            context['insumos_caducados'] = models.LoteInsumo.objects.filter(fecha_caducidad__lt=alert_date).order_by('fecha_caducidad')
            context['insumos_proximos_a_caducar'] = models.LoteInsumo.objects.filter(
                fecha_caducidad__gte=alert_date,
                fecha_caducidad__lte=proximos_30_dias
            ).order_by('fecha_caducidad')
            
        if user.groups.filter(name='Dentista').exists():
            perfil = get_object_or_404(models.PerfilDentista, usuario=user)
            citas_dentista_periodo = citas_periodo.filter(dentista=perfil)
            context['citas_dentista_periodo'] = citas_dentista_periodo.count()
            context['pacientes_atendidos_periodo'] = citas_dentista_periodo.filter(estado__in=['ATN', 'COM']).count()
            
        if user.groups.filter(name='Recepcionista').exists():
            context['citas_por_confirmar_periodo'] = citas_periodo.filter(estado='PRO').count()
            context['pacientes_nuevos_periodo'] = pacientes_periodo.count()
            
        return context

class UsuarioListView(TenantLoginRequiredMixin, ListView):
    model = User
    template_name = 'core/usuario_list.html'
    context_object_name = 'usuarios'
    paginate_by = 15

    def get_queryset(self):
        try:
            # Usar select_related para obtener el perfil_dentista si existe
            qs = User.objects.prefetch_related('groups').select_related('perfil_dentista').all().order_by('first_name', 'last_name')
            logger.debug(f"UsuarioListView queryset count={qs.count()}")
            return qs
        except Exception as e:
            logger.exception("Error obteniendo listado de usuarios")
            raise

class UsuarioCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = User
    form_class = forms.UserForm
    template_name = 'core/usuario_form.html'
    success_url = reverse_lazy('core:usuario_list')
    success_message = "Usuario '%(username)s' creado con éxito."

    def get_context_data(self, **kwargs):
        """Inicializar object para evitar AttributeError"""
        if not hasattr(self, 'object'):
            self.object = None
        return super().get_context_data(**kwargs)

    def form_valid(self, form):
        user = form.save(commit=False)
        user.set_password('password123')
        user.save()
        form.save_m2m()
        
        # Corregir la obtención de grupos - usar el campo 'rol' del formulario
        rol_seleccionado = form.cleaned_data.get('rol')
        if rol_seleccionado and rol_seleccionado.name in ['Administrador', 'Dentista']:
            models.PerfilDentista.objects.create(
                usuario=user,
                nombre=user.first_name,
                apellido=user.last_name,
                email=user.email
            )
        messages.warning(self.request, "Se asignó la contraseña temporal 'password123'. El usuario debe cambiarla.")
        return super().form_valid(form)

class UsuarioUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    
    model = User  # Usa tu modelo de usuario si es personalizado
    form_class = forms.UserForm
    template_name = 'core/usuario_edit.html'
    success_url = reverse_lazy('core:usuario_list')
    success_message = "Usuario actualizado correctamente."

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Verificar si el usuario pertenece al grupo "Dentista"
        es_dentista = self.object.groups.filter(name="Dentista").exists()
        context['mostrar_perfil_dentista'] = es_dentista

        perfil_obj = None
        if es_dentista:
            try:
                # Si existe un perfil de dentista, lo usamos como instancia
                perfil_obj = self.object.perfil_dentista
                if 'perfil_dentista_form' not in context or context['perfil_dentista_form'] is None:
                    context['perfil_dentista_form'] = forms.PerfilDentistaForm(instance=perfil_obj)
            except self.model.perfil_dentista.RelatedObjectDoesNotExist:
                if 'perfil_dentista_form' not in context or context['perfil_dentista_form'] is None:
                    context['perfil_dentista_form'] = forms.PerfilDentistaForm()
        else:
            context['perfil_dentista_form'] = None

        context['perfil_dentista_obj'] = perfil_obj
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        perfil_dentista_form = None

        # Verificar si el usuario es dentista
        es_dentista = self.object.groups.filter(name="Dentista").exists()

        if es_dentista:
            try:
                perfil = self.object.perfil_dentista
                perfil_dentista_form = forms.PerfilDentistaForm(request.POST, request.FILES, instance=perfil)
            except self.model.perfil_dentista.RelatedObjectDoesNotExist:
                perfil_dentista_form = forms.PerfilDentistaForm(request.POST, request.FILES)

            if perfil_dentista_form and not perfil_dentista_form.is_valid():
                # Volver a renderizar incluyendo errores del perfil
                context = self.get_context_data(perfil_dentista_form=perfil_dentista_form)
                context['form'] = form
                return self.render_to_response(context)

        if form.is_valid():
            # Guardar el formulario principal (UserForm)
            user = form.save()

            # Si es dentista y hay un formulario de perfil, guardarlo
            if es_dentista and perfil_dentista_form:
                perfil_dentista = perfil_dentista_form.save(commit=False)
                perfil_dentista.usuario = user
                perfil_dentista.save()
                perfil_dentista_form.save_m2m()  # Guardar especialidades (ManyToMany)

            return self.form_valid(form)
        # Si el form principal no es válido, volver a mostrar
        context = self.get_context_data(perfil_dentista_form=perfil_dentista_form)
        context['form'] = form
        return self.render_to_response(context)

class UsuarioDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = User
    template_name = 'core/usuario_confirm_delete.html'
    success_url = reverse_lazy('core:usuario_list')
    context_object_name = 'usuario'
    success_message = "Usuario eliminado con éxito."

    def dispatch(self, request, *args, **kwargs):
        # Solo administradores pueden eliminar usuarios
        if not (request.user.groups.filter(name='Administrador').exists() or request.user.is_superuser):
            messages.error(request, 'No tienes permisos para eliminar usuarios.')
            return redirect('core:usuario_list')
        
        # No permitir que el usuario se elimine a sí mismo
        usuario_a_eliminar = self.get_object()
        if usuario_a_eliminar == request.user:
            messages.error(request, 'No puedes eliminar tu propia cuenta.')
            return redirect('core:usuario_list')
        
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        usuario_nombre = f"{self.object.first_name} {self.object.last_name} ({self.object.username})"
        messages.success(self.request, f"Usuario '{usuario_nombre}' eliminado con éxito.")
        return super().form_valid(form)

class PacienteListView(TenantLoginRequiredMixin, ListView):
    model = models.Paciente
    template_name = 'core/paciente_list.html'
    context_object_name = 'pacientes'
    paginate_by = 20
    
    def get_queryset(self):
        # Optimizar consultas con select_related y prefetch_related
        queryset = models.Paciente.objects.select_related('usuario').prefetch_related(
            'respuestas_historial',
            'historial_clinico'
        ).all()
        
        form = forms.PacienteFiltroForm(self.request.GET or None)
        
        if form.is_valid():
            # Filtro de búsqueda
            busqueda = form.cleaned_data.get('busqueda')
            if busqueda:
                queryset = queryset.filter(
                    Q(nombre__icontains=busqueda) |
                    Q(apellido__icontains=busqueda) |
                    Q(email__icontains=busqueda) |
                    Q(telefono__icontains=busqueda)
                )
            
            # Filtro por estado del historial
            estado_historial = form.cleaned_data.get('estado_historial')
            if estado_historial == 'completo':
                # Pacientes que tienen historial clínico
                queryset = queryset.filter(
                    Q(respuestas_historial__isnull=False) | 
                    Q(historial_clinico__isnull=False)
                ).distinct()
            elif estado_historial == 'pendiente':
                # Pacientes sin historial clínico
                queryset = queryset.filter(
                    respuestas_historial__isnull=True,
                    historial_clinico__isnull=True
                )
            elif estado_historial == 'actualizar':
                # Pacientes con historial pero con más de 3 meses
                from datetime import timedelta
                tres_meses_atras = timezone.now() - timedelta(days=90)
                queryset = queryset.filter(
                    Q(respuestas_historial__isnull=False) | 
                    Q(historial_clinico__isnull=False),
                    Q(historial_clinico__fecha_evento__lt=tres_meses_atras) |
                    Q(historial_clinico__isnull=True, respuestas_historial__isnull=False)
                ).distinct()
            
            # Filtro por saldo pendiente
            if form.cleaned_data.get('con_saldo_pendiente'):
                queryset = queryset.filter(saldo_global__gt=0)
            
            # Filtro por acceso al portal
            if form.cleaned_data.get('con_acceso_portal'):
                queryset = queryset.filter(usuario__isnull=False)
            
            # Ordenamiento
            ordenar_por = form.cleaned_data.get('ordenar_por') or 'nombre'
            queryset = queryset.order_by(ordenar_por)
        else:
            queryset = queryset.order_by('nombre', 'apellido')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = forms.PacienteFiltroForm(self.request.GET or None)
        
        # Agregar información del estado del historial para cada paciente
        pacientes_con_estado = []
        for paciente in context['pacientes']:
            # Verificar si tiene respuestas al cuestionario
            tiene_respuestas = paciente.respuestas_historial.exists() if hasattr(paciente, 'respuestas_historial') else False
            
            # Verificar si tiene historial clínico
            historial_entries = paciente.historial_clinico.all() if hasattr(paciente, 'historial_clinico') else []
            tiene_historial = len(historial_entries) > 0
            
            # Determinar tipo de completado
            completado_portal = any('Auto-Completado por Paciente' in entry.descripcion_evento 
                                  for entry in historial_entries)
            completado_medico = any(entry.registrado_por is not None 
                                  for entry in historial_entries)
            
            # Calcular estado simplificado
            if tiene_respuestas or tiene_historial:
                # Verificar si necesita actualización (más de 3 meses)
                from datetime import timedelta
                tres_meses_atras = timezone.now() - timedelta(days=90)
                
                # Buscar la entrada más reciente del historial
                fecha_mas_reciente = None
                if historial_entries:
                    fecha_mas_reciente = max(entry.fecha_evento for entry in historial_entries)
                
                if fecha_mas_reciente and fecha_mas_reciente < tres_meses_atras:
                    estado_historial = 'actualizar'
                elif not historial_entries and tiene_respuestas:
                    # Solo tiene respuestas pero no historial clínico formal
                    estado_historial = 'actualizar'
                else:
                    estado_historial = 'completo'
            else:
                estado_historial = 'pendiente'
            
            paciente.estado_historial = estado_historial
            paciente.tiene_acceso_portal = bool(paciente.usuario)
            pacientes_con_estado.append(paciente)
        
        context['pacientes'] = pacientes_con_estado
        
        # Estadísticas rápidas
        total_pacientes = self.get_queryset().count()
        con_historial = self.get_queryset().filter(
            Q(respuestas_historial__isnull=False) | 
            Q(historial_clinico__isnull=False)
        ).distinct().count()
        con_portal = self.get_queryset().filter(usuario__isnull=False).count()
        con_saldo = self.get_queryset().filter(saldo_global__gt=0).count()
        
        context['estadisticas'] = {
            'total': total_pacientes,
            'con_historial': con_historial,
            'sin_historial': total_pacientes - con_historial,
            'con_portal': con_portal,
            'con_saldo': con_saldo,
        }
        
        return context

class PacienteDetailView(TenantLoginRequiredMixin, DetailView):
    model = models.Paciente
    template_name = 'core/paciente_detail.html'
    context_object_name = 'paciente'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            datos = models.DatosFiscales.objects.filter(paciente=self.object).first()
        except Exception:
            datos = None
        context['datos_fiscales'] = datos
        context['has_datos_fiscales'] = bool(datos)
        return context

class PacienteCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.Paciente
    template_name = 'core/paciente_form.html'
    form_class = forms.PacienteForm
    success_url = reverse_lazy('core:paciente_list')
    success_message = "Paciente '%(nombre)s %(apellido)s' creado con éxito."

class PacienteDatosFiscalesView(TenantLoginRequiredMixin, TemplateView):
    template_name = 'core/paciente_datos_fiscales_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        paciente = get_object_or_404(models.Paciente, pk=self.kwargs['pk'])
        context['paciente'] = paciente
        try:
            datos = models.DatosFiscales.objects.filter(paciente=paciente).first()
        except Exception:
            datos = None
        if self.request.POST:
            context['form'] = forms.DatosFiscalesForm(self.request.POST, instance=datos)
        else:
            context['form'] = forms.DatosFiscalesForm(instance=datos)
        return context

    def post(self, request, *args, **kwargs):
        paciente = get_object_or_404(models.Paciente, pk=self.kwargs['pk'])
        datos = models.DatosFiscales.objects.filter(paciente=paciente).first()
        form = forms.DatosFiscalesForm(request.POST, instance=datos)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.paciente = paciente
            obj.save()
            messages.success(request, 'Datos fiscales guardados correctamente.')
            return redirect('core:paciente_detail', pk=paciente.pk)
        messages.error(request, 'Por favor corrige los errores.')
        return self.render_to_response({'form': form, 'paciente': paciente})

class PacienteUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.Paciente
    template_name = 'core/paciente_form.html'
    form_class = forms.PacienteForm
    success_url = reverse_lazy('core:paciente_list')
    success_message = "Paciente '%(nombre)s %(apellido)s' actualizado con éxito."

class PacienteDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, DeleteView):
    model = models.Paciente
    template_name = 'core/paciente_confirm_delete.html'
    success_url = reverse_lazy('core:paciente_list')
    context_object_name = 'paciente'

    def form_valid(self, form):
        messages.success(self.request, f"Paciente '{self.object.nombre} {self.object.apellido}' eliminado con éxito.")
        return super().form_valid(form)

class ServicioListView(TenantLoginRequiredMixin, ListView):
    model = models.Servicio
    template_name = 'core/service_list.html'
    context_object_name = 'servicios'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = models.Servicio.objects.select_related('especialidad').filter(activo=True)
        
        # Filtros por parámetros GET
        busqueda = self.request.GET.get('busqueda')
        if busqueda:
            queryset = queryset.filter(
                Q(nombre__icontains=busqueda) | 
                Q(descripcion__icontains=busqueda) |
                Q(especialidad__nombre__icontains=busqueda)
            )
        
        especialidad = self.request.GET.get('especialidad')
        if especialidad:
            queryset = queryset.filter(especialidad_id=especialidad)
        
        precio_min = self.request.GET.get('precio_min')
        if precio_min:
            try:
                queryset = queryset.filter(precio__gte=float(precio_min))
            except ValueError:
                pass
        
        precio_max = self.request.GET.get('precio_max')
        if precio_max:
            try:
                queryset = queryset.filter(precio__lte=float(precio_max))
            except ValueError:
                pass
        
        duracion_min = self.request.GET.get('duracion_min')
        if duracion_min:
            try:
                queryset = queryset.filter(duracion_minutos__gte=int(duracion_min))
            except ValueError:
                pass
        
        duracion_max = self.request.GET.get('duracion_max')
        if duracion_max:
            try:
                queryset = queryset.filter(duracion_minutos__lte=int(duracion_max))
            except ValueError:
                pass
        
        # Ordenamiento
        orden = self.request.GET.get('orden', 'especialidad__nombre')
        if orden in ['nombre', 'precio', 'duracion_minutos', 'especialidad__nombre', '-precio', '-duracion_minutos']:
            queryset = queryset.order_by(orden, 'nombre')
        else:
            queryset = queryset.order_by('especialidad__nombre', 'nombre')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # === ANÁLISIS Y MÉTRICAS ===
        try:
            # Métricas básicas
            total_servicios = models.Servicio.objects.filter(activo=True).count()
            servicios_inactivos = models.Servicio.objects.filter(activo=False).count()
            
            # Precio promedio
            precio_promedio = models.Servicio.objects.filter(activo=True).aggregate(
                promedio=Avg('precio')
            )['promedio'] or 0
            
            # Duración promedio
            duracion_promedio = models.Servicio.objects.filter(activo=True).aggregate(
                promedio=Avg('duracion_minutos')
            )['promedio'] or 0
            
            # Servicios por especialidad
            servicios_por_especialidad = models.Especialidad.objects.annotate(
                total_servicios=Count('servicio', filter=Q(servicio__activo=True))
            ).filter(total_servicios__gt=0).order_by('-total_servicios')
            
            # Servicios más caros y más baratos
            servicio_mas_caro = models.Servicio.objects.filter(activo=True).order_by('-precio').first()
            servicio_mas_barato = models.Servicio.objects.filter(activo=True).order_by('precio').first()
            
            # Análisis de citas (servicios más utilizados en los últimos 30 días)
            hace_30_dias = timezone.now() - timedelta(days=30)
            servicios_populares = models.ServicioCita.objects.filter(
                cita__fecha_hora__gte=hace_30_dias
            ).values(
                'servicio__nombre', 
                'servicio__precio',
                'servicio__especialidad__nombre'
            ).annotate(
                veces_usado=Count('id'),
                ingresos_generados=Sum('precio_final')
            ).order_by('-veces_usado')[:5]
            
            # Análisis de rentabilidad (últimos 3 meses)
            hace_3_meses = timezone.now() - timedelta(days=90)
            servicios_rentables = models.ServicioCita.objects.filter(
                cita__fecha_hora__gte=hace_3_meses,
                cita__pagada=True
            ).values(
                'servicio__nombre',
                'servicio__precio', 
                'servicio__especialidad__nombre'
            ).annotate(
                total_ingresos=Sum('precio_final'),
                cantidad_realizados=Count('id')
            ).order_by('-total_ingresos')[:5]
            
            # Distribución de precios
            rangos_precios = {
                'bajo': models.Servicio.objects.filter(activo=True, precio__lt=500).count(),
                'medio': models.Servicio.objects.filter(activo=True, precio__gte=500, precio__lt=2000).count(),
                'alto': models.Servicio.objects.filter(activo=True, precio__gte=2000).count(),
            }
            
            # Distribución de duración
            rangos_duracion = {
                'corto': models.Servicio.objects.filter(activo=True, duracion_minutos__lt=30).count(),
                'medio': models.Servicio.objects.filter(activo=True, duracion_minutos__gte=30, duracion_minutos__lt=90).count(),
                'largo': models.Servicio.objects.filter(activo=True, duracion_minutos__gte=90).count(),
            }
            
            context.update({
                # Métricas básicas
                'total_servicios': total_servicios,
                'servicios_inactivos': servicios_inactivos,
                'precio_promedio': precio_promedio,
                'duracion_promedio': duracion_promedio,
                
                # Análisis
                'servicios_por_especialidad': servicios_por_especialidad,
                'servicio_mas_caro': servicio_mas_caro,
                'servicio_mas_barato': servicio_mas_barato,
                'servicios_populares': servicios_populares,
                'servicios_rentables': servicios_rentables,
                
                # Distribuciones
                'rangos_precios': rangos_precios,
                'rangos_duracion': rangos_duracion,
            })
            
        except Exception as e:
            logger.error(f"Error calculando métricas de servicios: {e}")
            # Valores por defecto en caso de error
            context.update({
                'total_servicios': 0,
                'servicios_inactivos': 0,
                'precio_promedio': 0,
                'duracion_promedio': 0,
                'servicios_por_especialidad': [],
                'servicios_populares': [],
                'servicios_rentables': [],
                'rangos_precios': {'bajo': 0, 'medio': 0, 'alto': 0},
                'rangos_duracion': {'corto': 0, 'medio': 0, 'largo': 0},
            })
        
        # === DATOS PARA FILTROS ===
        context['especialidades'] = models.Especialidad.objects.annotate(
            servicios_count=Count('servicio', filter=Q(servicio__activo=True))
        ).filter(servicios_count__gt=0).order_by('nombre')
        
        # Rangos de precios para filtros
        precios = models.Servicio.objects.filter(activo=True).aggregate(
            min_precio=Min('precio'),
            max_precio=Max('precio')
        )
        context['precio_min_disponible'] = precios['min_precio'] or 0
        context['precio_max_disponible'] = precios['max_precio'] or 0
        
        # Rangos de duración para filtros
        duraciones = models.Servicio.objects.filter(activo=True).aggregate(
            min_duracion=Min('duracion_minutos'),
            max_duracion=Max('duracion_minutos')
        )
        context['duracion_min_disponible'] = duraciones['min_duracion'] or 0
        context['duracion_max_disponible'] = duraciones['max_duracion'] or 0
        
        return context

class ServicioCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.Servicio
    form_class = forms.ServicioForm
    template_name = 'core/service_form.html'
    success_url = reverse_lazy('core:service_list')
    success_message = "Servicio '%(nombre)s' creado con éxito."

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = forms.ServicioInsumoFormSet(self.request.POST, prefix='insumos')
        else:
            context['formset'] = forms.ServicioInsumoFormSet(prefix='insumos')
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            return super().form_valid(form)
        else:
            return self.form_invalid(form)

class ServicioUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.Servicio
    form_class = forms.ServicioForm
    template_name = 'core/service_form.html'
    success_url = reverse_lazy('core:service_list')
    success_message = "Servicio '%(nombre)s' actualizado con éxito."

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = forms.ServicioInsumoFormSet(self.request.POST, instance=self.object, prefix='insumos')
        else:
            context['formset'] = forms.ServicioInsumoFormSet(instance=self.object, prefix='insumos')
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            return super().form_valid(form)
        else:
            return self.form_invalid(form)

class ServicioDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, DeleteView):
    model = models.Servicio
    template_name = 'core/service_confirm_delete.html'
    success_url = reverse_lazy('core:service_list')
    context_object_name = 'servicio'

    def form_valid(self, form):
        messages.success(self.request, f"Servicio '{self.object.nombre}' eliminado con éxito.")
        return super().form_valid(form)

class EspecialidadListView(TenantLoginRequiredMixin, ListView):
    model = models.Especialidad
    template_name = 'core/especialidad_list.html'
    context_object_name = 'especialidades'
    paginate_by = 10

class EspecialidadCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.Especialidad
    template_name = 'core/especialidad_form.html'
    fields = ['nombre']
    success_url = reverse_lazy('core:especialidad_list')
    success_message = "Especialidad '%(nombre)s' creada con éxito."

class EspecialidadUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.Especialidad
    template_name = 'core/especialidad_form.html'
    fields = ['nombre']
    success_url = reverse_lazy('core:especialidad_list')
    success_message = "Especialidad '%(nombre)s' actualizada con éxito."

class EspecialidadDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, DeleteView):
    model = models.Especialidad
    template_name = 'core/especialidad_confirm_delete.html'
    success_url = reverse_lazy('core:especialidad_list')
    context_object_name = 'especialidad'

    def form_valid(self, form):
        messages.success(self.request, f"Especialidad '{self.object.nombre}' eliminada con éxito.")
        return super().form_valid(form)

# REEMPLAZA tu PagoCreateView existente por:
class PagoCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.Pago  # ← Corregido: era 'mode'
    form_class = forms.PagoForm  # ✅ Este formulario existe
    template_name = 'core/pago_form.html'
    success_url = reverse_lazy('core:pago_list')
    success_message = "Pago creado exitosamente."
    
class PagoListView(TenantLoginRequiredMixin, ListView):
    model = models.Pago
    template_name = 'core/pago_list.html'
    context_object_name = 'pagos'
    paginate_by = 20

    def get_queryset(self):
        return models.Pago.objects.select_related('cita__paciente', 'paciente').order_by('-fecha_pago')

class PagoUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.Pago
    form_class = forms.PagoForm
    template_name = 'core/pago_form.html'
    success_url = reverse_lazy('core:pago_list')
    success_message = "Pago actualizado con éxito."

class PagoDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, DeleteView):
    model = models.Pago
    template_name = 'core/pago_confirm_delete.html'
    success_url = reverse_lazy('core:pago_list')
    context_object_name = 'pago'

    def form_valid(self, form):
        messages.success(self.request, f"Pago de {self.object.monto} eliminado con éxito.")
        return super().form_valid(form)

class CompraListView(TenantLoginRequiredMixin, ListView):
    model = models.Compra
    template_name = 'core/compra_list.html'
    context_object_name = 'compras'
    paginate_by = 15

    def get_queryset(self):
        return models.Compra.objects.select_related('proveedor').order_by('-fecha_compra')

class CompraCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.Compra
    form_class = forms.CompraForm
    template_name = 'core/compra_form.html'
    success_url = reverse_lazy('core:compra_list')
    success_message = "Compra creada con éxito."

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = forms.DetalleCompraFormSet(self.request.POST, prefix='detalles')
        else:
            context['formset'] = forms.DetalleCompraFormSet(prefix='detalles')
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if formset.is_valid():
            self.object = form.save(commit=False)
            
            total = 0
            for detalle_form in formset.cleaned_data:
                if detalle_form and not detalle_form.get('DELETE'):
                    cantidad = detalle_form.get('cantidad', 0)
                    precio = detalle_form.get('precio_unitario', 0)
                    total += cantidad * precio
            self.object.total = total

            self.object.save()
            formset.instance = self.object
            formset.save()
            return super().form_valid(form)
        else:
            return self.form_invalid(form)

class CompraUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.Compra
    form_class = forms.CompraForm
    template_name = 'core/compra_form.html'
    success_url = reverse_lazy('core:compra_list')
    success_message = "Compra actualizada con éxito."

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = forms.DetalleCompraFormSet(self.request.POST, instance=self.object, prefix='detalles')
        else:
            context['formset'] = forms.DetalleCompraFormSet(instance=self.object, prefix='detalles')
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if formset.is_valid():
            self.object = form.save(commit=False)
            
            total = 0
            for detalle_form in formset.cleaned_data:
                if detalle_form and not detalle_form.get('DELETE'):
                    cantidad = detalle_form.get('cantidad', 0)
                    precio = detalle_form.get('precio_unitario', 0)
                    total += cantidad * precio
            self.object.total = total

            self.object.save()
            formset.instance = self.object
            formset.save()
            return super().form_valid(form)
        else:
            return self.form_invalid(form)

class CompraDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, DeleteView):
    model = models.Compra
    template_name = 'core/compra_confirm_delete.html'
    success_url = reverse_lazy('core:compra_list')
    context_object_name = 'compra'

    def form_valid(self, form):
        messages.success(self.request, f"Compra a '{self.object.proveedor.nombre}' eliminada con éxito.")
        return super().form_valid(form)

class RecibirCompraView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.Compra
    template_name = 'core/compra_recibir.html'
    fields = []
    success_url = reverse_lazy('core:compra_list')
    success_message = "La compra ha sido marcada como 'Recibida' y el stock ha sido actualizado."

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = forms.DetalleCompraRecepcionFormSet(self.request.POST, instance=self.object, prefix='detalles')
        else:
            context['formset'] = forms.DetalleCompraRecepcionFormSet(instance=self.object, prefix='detalles')
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if formset.is_valid():
            with transaction.atomic():
                for detalle_form in formset:
                    detalle = detalle_form.instance
                    unidad_dental = detalle_form.cleaned_data.get('unidad_dental')
                    
                    if not unidad_dental:
                        messages.error(self.request, f"Debe seleccionar una unidad de destino para {detalle.insumo.nombre}.")
                        return self.form_invalid(form)

                    lote, created = models.LoteInsumo.objects.get_or_create(
                        insumo=detalle.insumo,
                        unidad_dental=unidad_dental,
                        numero_lote=detalle_form.cleaned_data.get('numero_lote'),
                        fecha_caducidad=detalle_form.cleaned_data.get('fecha_caducidad'),
                        defaults={'cantidad': detalle.cantidad}
                    )
                    if not created:
                        lote.cantidad += detalle.cantidad
                        lote.save()

                self.object.estado = 'RECIBIDA'
                self.object.save()
            
            return super().form_valid(form)
        else:
            return self.form_invalid(form)

class AgendaView(TenantLoginRequiredMixin, TemplateView):
    template_name = 'core/agenda.html'

    def get_context_data(self, **kwargs):
        try:
            context = super().get_context_data(**kwargs)
            context['cita_form'] = forms.CitaForm()
            context['paciente_form'] = forms.PacienteForm()
            context['dentistas'] = models.PerfilDentista.objects.filter(activo=True)
            logger.debug(f"AgendaView context dentistas={context['dentistas'].count()}")
            return context
        except Exception:
            logger.exception("Error construyendo el contexto de AgendaView")
            raise

class AgendaLegacyView(TenantLoginRequiredMixin, TemplateView):
    """Vista de calendario simplificado para dispositivos antiguos como iPad iOS 9.3"""
    template_name = 'core/agenda_legacy.html'

    def get_context_data(self, **kwargs):
        try:
            context = super().get_context_data(**kwargs)
            context['dentistas'] = models.PerfilDentista.objects.filter(activo=True)
            logger.debug(f"AgendaLegacyView context dentistas={context['dentistas'].count()}")
            return context
        except Exception:
            logger.exception("Error construyendo el contexto de AgendaLegacyView")
            raise

@tenant_login_required
def pacientes_api(request):
    """API para obtener lista de pacientes para el calendario legacy"""
    try:
        pacientes = models.Cliente.objects.all().order_by('nombre', 'apellido')[:100]  # Limitar a 100 para rendimiento
        pacientes_data = [{
            'id': p.id,
            'nombre': p.nombre,
            'apellido': p.apellido,
            'nombre_completo': f"{p.nombre} {p.apellido}"
        } for p in pacientes]
        
        return JsonResponse(pacientes_data, safe=False)
    except Exception as e:
        logger.exception("Error en pacientes_api")
        return JsonResponse({'error': str(e)}, status=500)

class CitaListView(TenantLoginRequiredMixin, ListView):
    model = models.Cita
    template_name = 'core/cita_list.html'
    context_object_name = 'citas'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = models.Cita.objects.select_related(
            'paciente', 'dentista', 'unidad_dental'
        ).prefetch_related(
            'servicios_planeados', 'servicios_realizados', 'pagos'
        ).order_by('-fecha_hora')
        
        user = self.request.user
        
        # FILTRO POR ROL: Si es dentista, solo ver sus citas
        if user.groups.filter(name='Dentista').exists():
            try:
                perfil_dentista = models.PerfilDentista.objects.get(usuario=user)
                queryset = queryset.filter(dentista=perfil_dentista)
            except models.PerfilDentista.DoesNotExist:
                messages.warning(self.request, 'No tienes un perfil de dentista asignado.')
                queryset = queryset.none()
        
        # FILTROS ADICIONALES
        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        
        dentista_id = self.request.GET.get('dentista')
        if dentista_id and user.groups.filter(name__in=['Administrador', 'Recepcionista']).exists():
            queryset = queryset.filter(dentista_id=dentista_id)
        
        fecha = self.request.GET.get('fecha')
        if fecha:
            queryset = queryset.filter(fecha_hora__date=fecha)
        
        # Filtro por paciente
        paciente = self.request.GET.get('paciente')
        if paciente:
            queryset = queryset.filter(
                models.Q(paciente__nombre__icontains=paciente) |
                models.Q(paciente__apellido__icontains=paciente)
            )
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Estados disponibles
        context['estados'] = [
            ('PRO', 'Programada'),
            ('CON', 'Confirmada'),
            ('ATN', 'Atendida'),
            ('COM', 'Completada'),
            ('CAN', 'Cancelada'),
        ]
        
        # Solo admin y recepción pueden filtrar por dentista
        context['puede_filtrar_dentista'] = user.groups.filter(
            name__in=['Administrador', 'Recepcionista']
        ).exists()
        
        if context['puede_filtrar_dentista']:
            context['dentistas'] = models.PerfilDentista.objects.filter(activo=True)
        
        # Estadísticas del queryset filtrado
        queryset = self.get_queryset()
        context['stats'] = {
            'programadas': queryset.filter(estado='PRO').count(),
            'confirmadas': queryset.filter(estado='CON').count(),
            'atendidas': queryset.filter(estado='ATN').count(),
            'completadas': queryset.filter(estado='COM').count(),
            'canceladas': queryset.filter(estado='CAN').count(),
            'total': queryset.count(),
        }
        
        # Añadir datos calculados a cada cita
        for cita in context['citas']:
            # Calcular costos
            cita.costo_estimado_calc = sum(s.precio for s in cita.servicios_planeados.all())
            cita.costo_real_calc = sum(s.precio for s in cita.servicios_realizados.all()) 
            cita.total_pagado_calc = cita.pagos.aggregate(total=Sum('monto'))['total'] or 0
            cita.saldo_pendiente_calc = cita.costo_real_calc - cita.total_pagado_calc
            
            # Calcular duración estimada  
            cita.duracion_estimada_calc = sum(s.duracion_minutos for s in cita.servicios_planeados.all())
        
        # Filtros aplicados (para mantener en los links de paginación)
        context['filtros_actuales'] = {
            'estado': self.request.GET.get('estado', ''),
            'dentista': self.request.GET.get('dentista', ''),
            'fecha': self.request.GET.get('fecha', ''),
            'paciente': self.request.GET.get('paciente', ''),
        }
        
        return context
    

class CambiarEstadoCitaView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, UpdateView):
    model = models.Cita
    fields = ['estado'] # Solo necesitamos el estado

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        new_status = request.POST.get('estado')
        
        if new_status not in [choice[0] for choice in models.Cita.ESTADOS_CITA]:
            return JsonResponse({'success': False, 'error': 'Estado no válido.'}, status=400)
            
        self.object.estado = new_status
        self.object.save()
        
        return JsonResponse({
            'success': True, 
            'message': f'Estado de la cita actualizado a "{self.object.get_estado_display()}"'
        })

    def http_method_not_allowed(self, request, *args, **kwargs):
        # Deshabilitar GET para esta vista
        return JsonResponse({'error': 'Método no permitido'}, status=405)

class CitasPendientesPagoListView(TenantLoginRequiredMixin, ListView):
    model = models.Cita
    template_name = 'core/citas_pendientes_pago.html'
    context_object_name = 'citas'
    paginate_by = 20

    def get_queryset(self):
        # Filtrar citas atendidas (ATN) o completadas (COM) con servicios realizados
        queryset = models.Cita.objects.filter(
            estado__in=['ATN', 'COM'],
            servicios_realizados__isnull=False
        ).select_related('paciente', 'dentista', 'unidad_dental').prefetch_related(
            'servicios_realizados', 'pagos'
        ).distinct().order_by('-fecha_hora')
        
        # Calcular saldo pendiente y filtrar solo las que tienen saldo > 0
        citas_con_saldo = []
        for cita in queryset:
            costo_real = sum(s.precio for s in cita.servicios_realizados.all())
            total_pagado = cita.pagos.aggregate(total=Sum('monto'))['total'] or 0
            saldo_pendiente = costo_real - total_pagado
            
            if saldo_pendiente > 0:
                # Añadir propiedades calculadas para usar en el template
                cita.costo_real_calc = costo_real
                cita.total_pagado_calc = total_pagado
                cita.saldo_pendiente_calc = saldo_pendiente
                citas_con_saldo.append(cita)
        
        # Aplicar filtros adicionales
        dentista_id = self.request.GET.get('dentista')
        if dentista_id:
            try:
                dentista_id_int = int(dentista_id)
                citas_con_saldo = [c for c in citas_con_saldo if c.dentista_id == dentista_id_int]
            except ValueError:
                pass
        
        fecha = self.request.GET.get('fecha')
        if fecha:
            from datetime import datetime
            try:
                fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
                citas_con_saldo = [c for c in citas_con_saldo if c.fecha_hora.date() == fecha_obj]
            except ValueError:
                pass
        
        paciente_nombre = self.request.GET.get('paciente')
        if paciente_nombre:
            citas_con_saldo = [
                c for c in citas_con_saldo 
                if paciente_nombre.lower() in f"{c.paciente.nombre} {c.paciente.apellido}".lower()
            ]
        
        return citas_con_saldo
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Solo admin y recepción pueden filtrar por dentista
        context['puede_filtrar_dentista'] = user.groups.filter(
            name__in=['Administrador', 'Recepcionista']
        ).exists()
        
        if context['puede_filtrar_dentista']:
            context['dentistas'] = models.PerfilDentista.objects.filter(activo=True)
        
        # Filtros aplicados (para mantener en los links de paginación)
        context['filtros_actuales'] = {
            'dentista': self.request.GET.get('dentista', ''),
            'fecha': self.request.GET.get('fecha', ''),
            'paciente': self.request.GET.get('paciente', ''),
        }
        
        # Estadísticas
        citas_list = context.get('citas', [])
        total_saldo_pendiente = sum(getattr(cita, 'saldo_pendiente_calc', 0) for cita in citas_list)
        context['total_saldo_pendiente'] = total_saldo_pendiente
        context['total_citas_pendientes'] = len(citas_list)
        
        return context
        
class FinalizarCitaView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.Cita
    form_class = forms.FinalizarCitaForm
    template_name = 'core/cita_finalizar.html'
    success_url = reverse_lazy('core:agenda')
    success_message = "La cita ha sido marcada como 'Atendida' y enviada a caja."

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['paciente_form'] = forms.PacientePlanPagoForm(self.request.POST, instance=self.object.paciente)
        else:
            context['paciente_form'] = forms.PacientePlanPagoForm(instance=self.object.paciente)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        paciente_form = context['paciente_form']
        
        if paciente_form.is_valid():
            with transaction.atomic():
                paciente = paciente_form.save()
                
                cita = form.save(commit=False)
                
                # Determinar dentista registrador (usuario autenticado o dentista de la cita)
                dentista_reg = getattr(getattr(self.request.user, 'perfil_dentista', None), 'pk', None)
                if dentista_reg:
                    dentista_reg = self.request.user.perfil_dentista
                else:
                    dentista_reg = cita.dentista
                
                servicios = form.cleaned_data['servicios_realizados']
                consumo_descripcion = []

                for servicio in servicios:
                    for item in servicio.servicioinsumo_set.all():
                        insumo_a_consumir = item.insumo
                        cantidad_necesaria = item.cantidad
                        
                        lotes_disponibles = models.LoteInsumo.objects.filter(
                            insumo=insumo_a_consumir,
                            unidad_dental=cita.unidad_dental,
                            cantidad__gt=0
                        ).order_by('fecha_caducidad')

                        for lote in lotes_disponibles:
                            if cantidad_necesaria <= 0:
                                break
                            
                            cantidad_a_tomar = min(lote.cantidad, cantidad_necesaria)
                            
                            lote.cantidad -= cantidad_a_tomar
                            lote.save()
                            
                            cantidad_necesaria -= cantidad_a_tomar
                            
                            consumo_descripcion.append(
                                f"{cantidad_a_tomar} de {insumo_a_consumir.nombre} (Lote: {lote.numero_lote or 'N/A'}, Cad: {lote.fecha_caducidad or 'N/A'})"
                            )
                
                if consumo_descripcion:
                    descripcion_evento = "Consumo de insumos para la cita: " + ", ".join(consumo_descripcion) + "."
                    models.HistorialClinico.objects.create(
                        paciente=cita.paciente,
                        descripcion_evento=descripcion_evento,
                        registrado_por=dentista_reg
                    )
                
                models.HistorialClinico.objects.create(
                    paciente=cita.paciente,
                    descripcion_evento=f"Cita atendida por Dr. {dentista_reg}. Servicios: {', '.join(s.nombre for s in servicios)}.",
                    registrado_por=dentista_reg
                )

                cita.estado = 'ATN'
                cita.save()
                form.save_m2m()
                
                paciente.actualizar_saldo_global()

            if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'redirect_url': str(self.get_success_url())}, status=200)
            return super().form_valid(form)
        else:
            if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors}, status=200)
            return self.form_invalid(form)

class HistorialPacienteView(TenantLoginRequiredMixin, DetailView):
    model = models.Paciente
    template_name = 'core/historial_paciente.html'
    context_object_name = 'paciente'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = forms.HistorialClinicoForm()

        import math
        
        def calcular_posiciones_arco(num_dientes, radio_x, radio_y, centro_x, centro_y, angulo_inicio, angulo_fin, es_inferior=False):
            dientes = []
            total_angulos = angulo_fin - angulo_inicio
            paso_angulo = total_angulos / (num_dientes - 1)
            
            for i in range(num_dientes):
                angulo_actual = angulo_inicio + (i * paso_angulo)
                angulo_rad = math.radians(angulo_actual)
                
                x = centro_x + radio_x * math.cos(angulo_rad)
                y = centro_y + radio_y * math.sin(angulo_rad)
                
                rotacion = angulo_actual + 90
                if es_inferior:
                    rotacion += 180

                dientes.append({'x': x, 'y': y, 'rotacion': rotacion})
            return dientes

        dientes_sup_data = [
            {'visual': 1, 'fdi': 18}, {'visual': 2, 'fdi': 17}, {'visual': 3, 'fdi': 16}, {'visual': 4, 'fdi': 15},
            {'visual': 5, 'fdi': 14}, {'visual': 6, 'fdi': 13}, {'visual': 7, 'fdi': 12}, {'visual': 8, 'fdi': 11},
            {'visual': 9, 'fdi': 21}, {'visual': 10, 'fdi': 22}, {'visual': 11, 'fdi': 23}, {'visual': 12, 'fdi': 24},
            {'visual': 13, 'fdi': 25}, {'visual': 14, 'fdi': 26}, {'visual': 15, 'fdi': 27}, {'visual': 16, 'fdi': 28},
        ]
        dientes_inf_data = [
            {'visual': 32, 'fdi': 48}, {'visual': 31, 'fdi': 47}, {'visual': 30, 'fdi': 46}, {'visual': 29, 'fdi': 45},
            {'visual': 28, 'fdi': 44}, {'visual': 27, 'fdi': 43}, {'visual': 26, 'fdi': 42}, {'visual': 25, 'fdi': 41},
            {'visual': 24, 'fdi': 31}, {'visual': 23, 'fdi': 32}, {'visual': 22, 'fdi': 33}, {'visual': 21, 'fdi': 34},
            {'visual': 20, 'fdi': 35}, {'visual': 19, 'fdi': 36}, {'visual': 18, 'fdi': 37}, {'visual': 17, 'fdi': 38},
        ]

        pos_sup = calcular_posiciones_arco(16, 420, 120, 460, 200, 195, 345)
        pos_inf = calcular_posiciones_arco(16, 420, 120, 460, 200, 165, 15, es_inferior=True)
        
        context['dientes_superiores'] = [{**d, **p} for d, p in zip(dientes_sup_data, pos_sup)]
        context['dientes_inferiores'] = [{**d, **p} for d, p in zip(dientes_inf_data, reversed(pos_inf))]

        return context

class HistorialClinicoCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, CreateView):
    model = models.HistorialClinico
    form_class = forms.HistorialClinicoForm

    def form_valid(self, form):
        paciente = models.Paciente.objects.get(pk=self.kwargs['cliente_id'])
        form.instance.paciente = paciente
        # Intentar con el dentista autenticado, si no, fallback a último dentista del paciente
        dentista = models.PerfilDentista.objects.filter(usuario=self.request.user).first()
        if dentista is None:
            ultima_cita = models.Cita.objects.filter(paciente=paciente).order_by('-fecha_hora').first()
            dentista = ultima_cita.dentista if ultima_cita else None
        form.instance.registrado_por = dentista
        
        dientes_seleccionados = form.cleaned_data['dientes_seleccionados'].split(',')
        diagnostico = form.cleaned_data['diagnostico']
        
        for diente_num in dientes_seleccionados:
            if diente_num:
                models.EstadoDiente.objects.update_or_create(
                    paciente=paciente,
                    numero_diente=int(diente_num),
                    defaults={'diagnostico': diagnostico}
                )
        
        messages.success(self.request, f"Se guardó la nota de evolución y se actualizaron {len(dientes_seleccionados)} dientes.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('core:paciente_history', kwargs={'pk': self.kwargs['cliente_id']})

@tenant_login_required
def agenda_events(request):
    from .timezone_utils import to_local_isoformat
    
    dentista_id = request.GET.get('dentista_id')
    citas = models.Cita.objects.exclude(estado='CAN')
    if dentista_id:
        citas = citas.filter(dentista_id=dentista_id)
    
    eventos = []
    for cita in citas:
        eventos.append({
            'id': cita.id,
            'title': f"{cita.paciente.nombre} {cita.paciente.apellido}",
            'start': to_local_isoformat(cita.fecha_hora),
            'extendedProps': {
                'estado': cita.estado,
                'notas': cita.notas,
                'paciente_id': cita.paciente.id,
                'dentista_id': cita.dentista.id if cita.dentista else None,
                'unidad_dental_id': cita.unidad_dental.id if cita.unidad_dental else None,
                'motivo': cita.motivo,
                'servicios': list(cita.servicios_planeados.values_list('id', flat=True)),
            }
        })
    
    return JsonResponse(eventos, safe=False)

@tenant_login_required
def get_horarios_ocupados(request):
    dentista_id = request.GET.get('dentista_id')
    fecha = request.GET.get('fecha')  # Formato: YYYY-MM-DD
    horarios_ocupados = []

    if dentista_id and fecha:
        fecha = datetime.strptime(fecha, '%Y-%m-%d')
        citas = models.Cita.objects.filter(
            dentista_id=dentista_id,
            fecha_hora__date=fecha,
            estado__in=['PEN', 'CON']  # Citas pendientes o confirmadas
        )
        for cita in citas:
            duracion = sum(s.duracion_minutos for s in cita.servicios_planeados.all())
            hora_fin = cita.fecha_hora + timedelta(minutes=duracion)
            horarios_ocupados.append({
                'inicio': cita.fecha_hora.isoformat(),
                'fin': hora_fin.isoformat()
            })

    return JsonResponse({'horarios_ocupados': horarios_ocupados})

@tenant_login_required
def get_horarios_disponibles_api(request, dentista_id):
    # Verificar si el usuario está autenticado para peticiones AJAX
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Autenticación requerida'}, status=401)
    
    fecha_str = request.GET.get('fecha')  # Formato: YYYY-MM-DD

    if not fecha_str:
        return JsonResponse({'error': 'Fecha requerida'}, status=400)

    try:
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        dia_semana = fecha.weekday()
        dentista = models.PerfilDentista.objects.get(pk=dentista_id)
        
        # Obtener horarios laborales
        horarios = models.HorarioLaboral.objects.filter(dentista=dentista, dia_semana=dia_semana, activo=True)
        if not horarios:
            return JsonResponse({'horarios_disponibles': []}, safe=False)

        # Calcular horarios ocupados
        citas = models.Cita.objects.filter(
            dentista=dentista,
            fecha_hora__date=fecha,
            estado__in=['PRO', 'CON']
        )
        ocupados = []
        for cita in citas:
            # Asegurarse de que cita.duracion_estimada no sea None o cause error
            duracion = cita.duracion_estimada if cita.duracion_estimada is not None else 0
            hora_fin = cita.fecha_hora + timedelta(minutes=duracion)
            ocupados.append((cita.fecha_hora, hora_fin))

        # Generar horarios disponibles
        disponibles = []
        intervalo = 30  # minutos
        
        now = timezone.now() # Obtener la hora actual con zona horaria
        today = now.date()

        for horario in horarios:
            # Hacer los datetimes aware para compararlos con cita.fecha_hora
            inicio = timezone.make_aware(datetime.combine(fecha, horario.hora_inicio))
            fin = timezone.make_aware(datetime.combine(fecha, horario.hora_fin))
            actual = inicio

            # Si la fecha es hoy, ajustar el inicio para que sea la hora actual o posterior
            if fecha == today:
                # Asegurarse de que `now` también sea aware para la comparación
                now_aware = timezone.make_aware(datetime.now())
                if actual < now_aware:
                    actual = now_aware
                    # Redondear al próximo intervalo si no coincide exactamente
                    if actual.minute % intervalo != 0:
                        actual = actual + timedelta(minutes=intervalo - (actual.minute % intervalo))
                    actual = actual.replace(second=0, microsecond=0)

            while actual + timedelta(minutes=intervalo) <= fin:
                es_disponible = True
                for ocup_start, ocup_end in ocupados:
                    if actual < ocup_end and (actual + timedelta(minutes=intervalo)) > ocup_start:
                        es_disponible = False
                        break
                if es_disponible:
                    disponibles.append(actual.strftime('%H:%M'))
                actual += timedelta(minutes=intervalo)

        return JsonResponse({'horarios_disponibles': disponibles}, safe=False)
    except models.PerfilDentista.DoesNotExist:
        return JsonResponse({'error': 'Dentista no encontrado'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc() # Imprimir el traceback completo
        return JsonResponse({'error': str(e)}, status=500)

@tenant_login_required
def cita_detail_api(request, pk):
    from .timezone_utils import to_local_strftime
    
    try:
        cita = get_object_or_404(models.Cita, pk=pk)
        servicios_planeados_ids = list(cita.servicios_planeados.values_list('id', flat=True))
        
        data = {
            'paciente': cita.paciente.id,
            'dentista': cita.dentista.id if cita.dentista else '',
            'fecha_hora': to_local_strftime(cita.fecha_hora),
            'motivo': cita.motivo,
            'estado': cita.estado,
            'servicios_planeados': servicios_planeados_ids,
            'unidad_dental': cita.unidad_dental.id if cita.unidad_dental else '',
        }
        return JsonResponse(data)
    except models.Cita.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Cita no encontrada'}, status=404)

class DentistaListView(TenantLoginRequiredMixin, ListView):
    model = models.PerfilDentista
    template_name = 'core/dentista_list.html'
    context_object_name = 'dentistas'

@tenant_login_required
def odontograma_api_get(request, cliente_id):
    """API para obtener el estado del odontograma completo de 48 dientes de un paciente"""
    try:
        cliente = models.Paciente.objects.get(pk=cliente_id)
        sano = models.Diagnostico.objects.filter(nombre='SANO').first()
        if not sano:
            # Crear diagnóstico SANO si no existe
            sano = models.Diagnostico.objects.create(nombre='SANO', color_hex='#FFFFFF', icono_svg='')
        
        # Lista completa de 48 dientes usando numeración FDI extendida
        # Dientes regulares (32 dientes estándar)
        dientes_regulares = (
            list(range(11, 19)) +  # Cuadrante I: 11-18
            list(range(21, 29)) +  # Cuadrante II: 21-28
            list(range(31, 39)) +  # Cuadrante III: 31-38
            list(range(41, 49))    # Cuadrante IV: 41-48
        )
        
        # Supernumerarios usando numeración extendida
        supernumerarios = [
            19, 110,    # Cuadrante I: distales
            29, 210,    # Cuadrante II: distales
            39, 310,    # Cuadrante III: distales
            49, 410     # Cuadrante IV: distales
        ]
        
        # Combinar todos los dientes (48 total)
        todos_los_dientes = dientes_regulares + supernumerarios
        
        # Inicializar todos los dientes como SANOS
        data = {}
        for diente in todos_los_dientes:
            data[diente] = {
                'diagnostico_id': sano.id,
                'diagnostico_nombre': sano.nombre,
                'diagnostico_color': sano.color_hex,
                'diagnostico_icono': sano.icono_svg,
                'color_seleccionado': '',
                'es_supernumerario': diente in supernumerarios,
                'cuadrante': _determinar_cuadrante(diente)
            }

        # Sobrescribir con estados guardados del paciente
        estados_guardados = cliente.odontograma.select_related('diagnostico')
        for estado in estados_guardados:
            if estado.numero_diente in data:  # Verificar que el diente esté en nuestra lista de 48
                data[estado.numero_diente].update({
                    'diagnostico_id': estado.diagnostico.id,
                    'diagnostico_nombre': estado.diagnostico.nombre,
                    'diagnostico_color': estado.diagnostico.color_hex,
                    'diagnostico_icono': estado.diagnostico.icono_svg,
                    'color_seleccionado': estado.color_seleccionado
                })
        
        return JsonResponse({
            'dientes': data,
            'total_dientes': len(todos_los_dientes),
            'regulares': len(dientes_regulares),
            'supernumerarios': len(supernumerarios)
        })

    except models.Paciente.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Paciente no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

def _determinar_cuadrante(numero_diente):
    """Determina el cuadrante dental basado en el número FDI"""
    if numero_diente in [19, 110] or (11 <= numero_diente <= 18):
        return 1
    elif numero_diente in [29, 210] or (21 <= numero_diente <= 28):
        return 2
    elif numero_diente in [39, 310] or (31 <= numero_diente <= 38):
        return 3
    elif numero_diente in [49, 410] or (41 <= numero_diente <= 48):
        return 4
    else:
        return 0  # No determinado

@tenant_login_required
def odontograma_48_view(request, cliente_id):
    """Vista principal para el odontograma completo de 48 dientes"""
    try:
        paciente = get_object_or_404(models.Paciente, pk=cliente_id)
        
        # Obtener diagnósticos disponibles
        diagnosticos = models.Diagnostico.objects.all().order_by('nombre')
        
        # Verificar permisos (opcional: agregar lógica de permisos aquí)
        
        context = {
            'paciente': paciente,
            'diagnosticos': diagnosticos,
            'es_odontograma_48': True,
        }
        
        return render(request, 'core/odontograma_48.html', context)
        
    except models.Paciente.DoesNotExist:
        messages.error(request, 'Paciente no encontrado')
        return redirect('core:paciente_list')
    except Exception as e:
        messages.error(request, f'Error al cargar el odontograma: {str(e)}')
        return redirect('core:paciente_list')

@tenant_login_required
@csrf_exempt
@require_POST
def odontograma_api_update(request, cliente_id):
    try:
        data = json.loads(request.body)
        numero_diente = data.get('numero_diente')
        diagnostico_id = data.get('diagnostico_id')
        color_seleccionado = data.get('color_seleccionado', '')

        paciente = models.Paciente.objects.get(pk=cliente_id)
        diagnostico = models.Diagnostico.objects.get(pk=diagnostico_id)
        dentista = models.PerfilDentista.objects.filter(usuario=request.user).first()
        if dentista is None:
            # Fallback: tomar el dentista de la última cita del paciente
            ultima_cita = models.Cita.objects.filter(paciente=paciente).order_by('-fecha_hora').first()
            dentista = ultima_cita.dentista if ultima_cita else None

        obj, created = models.EstadoDiente.objects.update_or_create(
            paciente=paciente,
            numero_diente=numero_diente,
            defaults={
                'diagnostico': diagnostico,
                'color_seleccionado': color_seleccionado
            }
        )

        descripcion = (
            f"Se {'asignó' if created else 'actualizó'} el diagnóstico '{diagnostico.nombre}' "
            f"al diente {numero_diente}."
        )
        if color_seleccionado:
            descripcion += f" Se asignó el color {color_seleccionado}."

        models.HistorialClinico.objects.create(
            paciente=paciente,
            descripcion_evento=descripcion,
            registrado_por=dentista
        )

        return JsonResponse({'status': 'success', 'message': 'Odontograma actualizado.'})

    except (models.Paciente.DoesNotExist, models.Diagnostico.DoesNotExist) as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@tenant_login_required
def diagnostico_api_list(request):
    diagnosticos = models.Diagnostico.objects.all().order_by('nombre')
    data = [{'id': d.id, 'nombre': d.nombre, 'color_hex': d.color_hex, 'icono_svg': d.icono_svg} for d in diagnosticos]
    return JsonResponse(data, safe=False)

@tenant_login_required
def reporte_ingresos_api(request):
    form = forms.ReporteIngresosForm(request.GET or None)
    data = {'labels': [], 'values': []}
    from datetime import timedelta
    today = timezone.now().date()
    default_inicio = today - timedelta(days=30)
    default_fin = today
    if form.is_valid():
        fecha_inicio = form.cleaned_data['fecha_inicio'] or default_inicio
        fecha_fin = form.cleaned_data['fecha_fin'] or default_fin
    else:
        fecha_inicio, fecha_fin = default_inicio, default_fin
    ingresos = models.Pago.objects.filter(
        fecha_pago__date__gte=fecha_inicio,
        fecha_pago__date__lte=fecha_fin
    ).values('fecha_pago__date').annotate(total_dia=Sum('monto')).order_by('fecha_pago__date')
    for item in ingresos:
        data['labels'].append(item['fecha_pago__date'].strftime('%d/%m/%Y'))
        data['values'].append(float(item['total_dia']))
    return JsonResponse(data)

@tenant_login_required
def reporte_saldos_api(request):
    # USAR SALDO_GLOBAL DE PACIENTES
    saldos = models.Paciente.objects.filter(
        saldo_global__gt=0
    ).values('nombre', 'apellido', 'saldo_global').order_by('-saldo_global')[:10]
    
    data = {
        'labels': [f"{item['nombre']} {item['apellido']}" for item in saldos],
        'values': [float(item['saldo_global']) for item in saldos]
    }
    return JsonResponse(data)

class ReciboPagoView(TenantLoginRequiredMixin, DetailView):
    model = models.Pago
    template_name = 'core/recibo_pago.html'
    context_object_name = 'pago'

@tenant_login_required
def generar_recibo_pdf(request, pk):
    pago = get_object_or_404(models.Pago, pk=pk)
    formato = request.GET.get('formato', 'carta')
    
    response = HttpResponse(content_type='application/pdf')
    filename = f"recibo_{pago.id}_{formato}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    
    if formato == 'ticket':
        width, height = 80 * mm, 200 * mm
        _generar_recibo_ticket(response, pago, width, height, request)  # ← Pasar request
    else:
        _generar_recibo_carta(response, pago, request)  # ← Pasar request
    
    return response

def _generar_recibo_carta(response, pago, request):  # ← Agregar request
    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    
    # OBTENER TENANT DESDE REQUEST
    tenant = request.tenant
    
    if tenant.logo:
        try:
            logo = Image(tenant.logo.path, width=50, height=50)
            logo.hAlign = 'LEFT'
            story.append(logo)
        except Exception:
            story.append(Paragraph(f"<h1>{tenant.nombre}</h1>", styles['h1']))
    else:
        story.append(Paragraph(f"<h1>{tenant.nombre}</h1>", styles['h1']))
        
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"<b>Recibo de Pago #{pago.id}</b>", styles['h2']))
    story.append(Paragraph(f"Fecha: {pago.fecha_pago.strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 24))
    story.append(Paragraph("<b>Paciente:</b>", styles['h4']))
    
    # Obtener paciente desde cita o directamente del pago
    paciente = pago.cita.paciente if pago.cita else pago.paciente
    story.append(Paragraph(f"{paciente.nombre} {paciente.apellido}", styles['Normal']))
    
    if paciente.email:
        story.append(Paragraph(f"{paciente.email}", styles['Normal']))
        
    story.append(Spacer(1, 24))
    story.append(Paragraph("<b>Detalles del Pago:</b>", styles['h4']))
    
    data = [['Descripción', 'Monto']]
    total_servicios = 0
    
    if pago.cita:
        # Pago por cita - mostrar servicios realizados
        for servicio in pago.cita.servicios_realizados.all():
            data.append([servicio.nombre, f"${servicio.precio:,.2f}"])
            total_servicios += servicio.precio
    else:
        # Abono general - mostrar concepto simple
        data.append(["Abono general", f"${pago.monto:,.2f}"])
        total_servicios = pago.monto
        
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ])
    
    tbl = Table(data, colWidths=[300, 100])
    tbl.setStyle(style)
    story.append(tbl)
    
    styles.add(ParagraphStyle(name='Right', alignment=TA_RIGHT))
    story.append(Spacer(1, 12))
    if pago.cita:
        story.append(Paragraph(f"<b>Total Servicios:</b> ${total_servicios:,.2f}", styles['Right']))
    else:
        story.append(Paragraph(f"<b>Concepto:</b> Abono general", styles['Right']))
    
    story.append(Paragraph(f"<b>Monto Pagado ({pago.metodo_pago}):</b> ${pago.monto:,.2f}", styles['Right']))
    
    # USAR SALDO_GLOBAL DEL PACIENTE
    paciente = pago.cita.paciente if pago.cita else pago.paciente
    story.append(Paragraph(f"<b>Saldo Pendiente del Paciente:</b> ${paciente.saldo_global:,.2f}", styles['Right']))
    
    doc.build(story)


def _generar_recibo_ticket(response, pago, width, height, request):  # ← Agregar request
    from reportlab.pdfgen import canvas
    
    c = canvas.Canvas(response, pagesize=(width, height))
    
    x_pos = 3 * mm
    y_pos = height - (10 * mm)
    line_height = 5 * mm
    
    # OBTENER TENANT DESDE REQUEST
    tenant = request.tenant
    
    if tenant.logo:
        try:
            c.drawImage(tenant.logo.path, x_pos, y_pos - 10*mm, width=20*mm, height=20*mm, preserveAspectRatio=True)
            y_pos -= 25 * mm
        except Exception:
            pass
    
    c.setFont("Helvetica-Bold", 12)
    c.drawString(x_pos, y_pos, tenant.nombre)
    y_pos -= line_height * 2
    c.setFont("Helvetica", 9)
    c.drawString(x_pos, y_pos, f"Recibo de Pago #{pago.id}")
    y_pos -= line_height
    c.drawString(x_pos, y_pos, f"Fecha: {pago.fecha_pago.strftime('%d/%m/%Y %H:%M')}")
    y_pos -= line_height * 1.5
    
    # Obtener paciente desde cita o directamente del pago
    paciente = pago.cita.paciente if pago.cita else pago.paciente
    c.drawString(x_pos, y_pos, f"Paciente: {paciente}")
    y_pos -= line_height * 2
    c.line(x_pos, y_pos, width - x_pos, y_pos)
    y_pos -= line_height
    c.setFont("Helvetica-Bold", 9)
    
    total_servicios = 0
    
    if pago.cita:
        # Pago por cita - mostrar servicios
        c.drawString(x_pos, y_pos, "Servicios:")
        y_pos -= line_height
        c.setFont("Helvetica", 8)
        
        for servicio in pago.cita.servicios_realizados.all():
            c.drawString(x_pos + 2*mm, y_pos, f"- {servicio.nombre}")
            c.drawRightString(width - x_pos, y_pos, f"${servicio.precio:,.2f}")
            total_servicios += servicio.precio
            y_pos -= line_height
        y_pos -= line_height
    else:
        # Abono general - mostrar concepto
        c.drawString(x_pos, y_pos, "Concepto:")
        y_pos -= line_height
        c.setFont("Helvetica", 8)
        c.drawString(x_pos + 2*mm, y_pos, "- Abono general")
        c.drawRightString(width - x_pos, y_pos, f"${pago.monto:,.2f}")
        total_servicios = pago.monto
        y_pos -= line_height * 2
    
    c.setFont("Helvetica-Bold", 9)
    if pago.cita:
        c.drawRightString(width - x_pos, y_pos, f"Total: ${total_servicios:,.2f}")
        y_pos -= line_height
    
    c.drawRightString(width - x_pos, y_pos, f"Pagado: ${pago.monto:,.2f}")
    y_pos -= line_height
    
    # USAR SALDO_GLOBAL DEL PACIENTE
    paciente = pago.cita.paciente if pago.cita else pago.paciente
    c.drawRightString(width - x_pos, y_pos, f"Saldo: ${paciente.saldo_global:,.2f}")
    y_pos -= line_height * 2
    c.setFont("Helvetica-Oblique", 8)
    c.drawCentredString(width / 2, y_pos, "Gracias por su preferencia")
    c.showPage()
    c.save()

class DashboardCofeprisView(TenantLoginRequiredMixin, TemplateView):
    template_name = 'core/cofepris/dashboard_cofepris.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from datetime import date, timedelta
        hoy = date.today()
        proximos_60_dias = hoy + timedelta(days=60)

        context['aviso_vencimiento'] = models.AvisoFuncionamiento.objects.filter(
            consultorio=self.request.tenant,
            fecha_vencimiento__lte=proximos_60_dias
        ).first()

        context['equipos_proximos_a_vencer'] = models.Equipo.objects.filter(
            unidad_trabajo__in=models.UnidadDental.objects.all(),
            fecha_vencimiento_calibracion__lte=proximos_60_dias
        ).order_by('fecha_vencimiento_calibracion')

        return context

class DiagnosticoListView(TenantLoginRequiredMixin, ListView):
    model = models.Diagnostico
    template_name = 'core/diagnostico_list.html'
    context_object_name = 'diagnosticos'

class DiagnosticoCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.Diagnostico
    template_name = 'core/diagnostico_form.html'
    fields = ['nombre', 'color_hex', 'icono_svg']
    success_url = reverse_lazy('core:diagnostico_list')
    success_message = "Diagnóstico '%(nombre)s' creado con éxito."

class DiagnosticoUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.Diagnostico
    template_name = 'core/diagnostico_form.html'
    fields = ['nombre', 'color_hex', 'icono_svg']
    success_url = reverse_lazy('core:diagnostico_list')
    success_message = "Diagnóstico '%(nombre)s' actualizado con éxito."

class DiagnosticoDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = models.Diagnostico
    template_name = 'core/diagnostico_confirm_delete.html'
    success_url = reverse_lazy('core:diagnostico_list')
    context_object_name = 'diagnostico'
    success_message = "Diagnóstico eliminado con éxito."

class AvisoFuncionamientoListView(TenantLoginRequiredMixin, ListView):
    model = models.AvisoFuncionamiento
    template_name = 'core/cofepris/aviso_funcionamiento_list.html'
    context_object_name = 'avisos'

    def get_queryset(self):
        return models.AvisoFuncionamiento.objects.filter(consultorio=self.request.tenant)

class AvisoFuncionamientoCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.AvisoFuncionamiento
    form_class = forms.AvisoFuncionamientoForm
    template_name = 'core/cofepris/aviso_funcionamiento_form.html'
    success_url = reverse_lazy('core:aviso_list')
    success_message = "Aviso de Funcionamiento creado con éxito."

    def form_valid(self, form):
        form.instance.consultorio = self.request.tenant
        return super().form_valid(form)

class AvisoFuncionamientoUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.AvisoFuncionamiento
    form_class = forms.AvisoFuncionamientoForm
    template_name = 'core/cofepris/aviso_funcionamiento_form.html'
    success_url = reverse_lazy('core:aviso_list')
    success_message = "Aviso de Funcionamiento actualizado con éxito."

class EquipoListView(TenantLoginRequiredMixin, ListView):
    model = models.Equipo
    template_name = 'core/cofepris/equipo_list.html'
    context_object_name = 'equipos'

class EquipoCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.Equipo
    form_class = forms.EquipoForm
    template_name = 'core/cofepris/equipo_form.html'
    success_url = reverse_lazy('core:equipo_list')
    success_message = "Equipo '%(nombre)s' registrado con éxito."

class EquipoUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.Equipo
    form_class = forms.EquipoForm
    template_name = 'core/cofepris/equipo_form.html'
    success_url = reverse_lazy('core:equipo_list')
    success_message = "Equipo '%(nombre)s' actualizado con éxito."

class EquipoDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = models.Equipo
    template_name = 'core/cofepris/equipo_confirm_delete.html'
    success_url = reverse_lazy('core:equipo_list')
    context_object_name = 'equipo'
    success_message = "Equipo eliminado con éxito."

class ResiduosListView(TenantLoginRequiredMixin, ListView):
    model = models.Residuos
    template_name = 'core/cofepris/residuos_list.html'
    context_object_name = 'recolecciones'

class ResiduosCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.Residuos
    form_class = forms.ResiduosForm
    template_name = 'core/cofepris/residuos_form.html'
    success_url = reverse_lazy('core:residuos_list')
    success_message = "Recolección de residuos registrada con éxito."

    def form_valid(self, form):
        form.instance.consultorio = self.request.tenant
        return super().form_valid(form)

class ResiduosUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.Residuos
    form_class = forms.ResiduosForm
    template_name = 'core/cofepris/residuos_form.html'
    success_url = reverse_lazy('core:residuos_list')
    success_message = "Recolección de residuos actualizada con éxito."

class ResiduosDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = models.Residuos
    template_name = 'core/cofepris/residuos_confirm_delete.html'
    success_url = reverse_lazy('core:residuos_list')
    context_object_name = 'recoleccion'
    success_message = "Registro de recolección eliminado con éxito."

@tenant_login_required
def exportar_ingresos_excel(request):
    form = forms.ReporteIngresosForm(request.GET or None)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_ingresos.xlsx"'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Ingresos'
    headers = ['Fecha de Pago', 'Paciente', 'Cita', 'Método de Pago', 'Monto']
    worksheet.append(headers)
    if form.is_valid():
        pagos = models.Pago.objects.filter(
            fecha_pago__date__gte=form.cleaned_data['fecha_inicio'],
            fecha_pago__date__lte=form.cleaned_data['fecha_fin']
        ).select_related('cita__paciente')
        for pago in pagos:
            worksheet.append([
                pago.fecha_pago.strftime('%Y-%m-%d %H:%M'),
                str(pago.cita.paciente),
                f"Cita #{pago.cita.id}",
                pago.metodo_pago,
                pago.monto
            ])
    workbook.save(response)
    return response

@tenant_login_required
def exportar_saldos_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_saldos.xlsx"'
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Saldos Pendientes'
    
    headers = ['Paciente', 'Saldo Pendiente']
    worksheet.append(headers)
    
    # USAR SALDO_GLOBAL DE PACIENTES
    pacientes = models.Paciente.objects.filter(saldo_global__gt=0).order_by('-saldo_global')
    for paciente in pacientes:
        worksheet.append([
            f"{paciente.nombre} {paciente.apellido}", 
            float(paciente.saldo_global)
        ])
    workbook.save(response)
    return response

@tenant_login_required
def exportar_facturacion_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_facturacion.xlsx"'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Facturación'
    headers = ['Fecha Cita', 'Paciente', 'RFC Receptor', 'Nombre Receptor', 'CP Receptor', 'Régimen Fiscal', 'Uso CFDI', 'Forma Pago (SAT)', 'Método Pago (SAT)', 'Servicios', 'Monto Pagado']
    worksheet.append(headers)
    citas = models.Cita.objects.filter(requiere_factura=True).select_related('paciente').prefetch_related('servicios_realizados', 'pagos')
    for cita in citas:
        servicios = ", ".join([s.nombre for s in cita.servicios_realizados.all()])
        monto_pagado = cita.pagos.aggregate(total=Sum('monto'))['total'] or 0
        pago = cita.pagos.order_by('-fecha_pago').first()
        df = getattr(cita.paciente, 'datos_fiscales', None)
        cp = getattr(df, 'codigo_postal', None) or getattr(cita.paciente, 'codigo_postal', None) or 'N/A'
        regimen = getattr(getattr(df, 'regimen_fiscal', None), 'codigo', 'N/A')
        uso = getattr(getattr(df, 'uso_cfdi', None), 'codigo', 'N/A')
        forma = getattr(getattr(pago, 'forma_pago_sat', None), 'codigo', 'N/A') if pago else 'N/A'
        metodo = getattr(getattr(pago, 'metodo_sat', None), 'codigo', 'N/A') if pago else 'N/A'
        worksheet.append([
            cita.fecha_hora.strftime('%Y-%m-%d'),
            str(cita.paciente),
            getattr(df, 'rfc', 'N/A'),
            getattr(df, 'razon_social', str(cita.paciente)),
            cp,
            regimen,
            uso,
            forma,
            metodo,
            servicios,
            monto_pagado
        ])
    workbook.save(response)
    return response

class InvitarPacienteView(TenantLoginRequiredMixin, TemplateView):
    template_name = 'core/paciente_invitar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['paciente'] = get_object_or_404(models.Paciente, pk=self.kwargs['pk'])
        return context

    def post(self, request, *args, **kwargs):
        paciente = get_object_or_404(models.Paciente, pk=self.kwargs['pk'])
        if paciente.usuario:
            messages.error(request, 'Este paciente ya tiene una cuenta de usuario.')
            return redirect('core:paciente_detail', pk=paciente.pk)

        username = paciente.email or f'{paciente.nombre.lower()}.{paciente.apellido.lower()}{paciente.pk}'
        if User.objects.filter(username=username).exists():
            username = f'{username}{random.randint(10,99)}'
        
        password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))

        user = User.objects.create_user(username=username, email=paciente.email, password=password)
        user.first_name = paciente.nombre
        user.last_name = paciente.apellido
        
        paciente_group, _ = Group.objects.get_or_create(name='Paciente')
        user.groups.add(paciente_group)
        user.save()

        paciente.usuario = user
        paciente.save()

        messages.success(request, f'Se creó la cuenta para {paciente}. Usuario: {username}, Contraseña: {password}')
        return redirect('core:paciente_detail', pk=paciente.pk)

class PacientePagosListView(TenantLoginRequiredMixin, ListView):
    model = models.Pago
    template_name = 'core/portal/pago_list.html'
    context_object_name = 'pagos'
    paginate_by = 10

    def get_queryset(self):
        return models.Pago.objects.filter(cita__paciente__usuario=self.request.user).order_by('-fecha_pago')

class PortalHistorialPacienteView(TenantLoginRequiredMixin, TemplateView):
    """Vista para que el paciente vea su historial clínico desde el portal"""
    template_name = 'core/portal/historial_paciente.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Verificar que el usuario sea un paciente
        if not hasattr(self.request.user, 'paciente_perfil'):
            raise PermissionDenied("Acceso no autorizado")
        
        paciente = self.request.user.paciente_perfil
        context['paciente'] = paciente
        
        # Obtener historial clínico existente
        context['historial_clinico'] = models.HistorialClinico.objects.filter(
            paciente=paciente
        ).order_by('-fecha_registro')
        
        # Verificar si ya tiene respuestas al cuestionario
        respuestas_existentes = models.RespuestaHistorial.objects.filter(
            paciente=paciente
        ).exists()
        context['tiene_cuestionario_completo'] = respuestas_existentes
        
        return context

class PortalCompletarHistorialView(TenantLoginRequiredMixin, TemplateView):
    """Vista para que el paciente complete su historial clínico mejorado desde el portal"""
    template_name = 'core/portal/completar_historial.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Verificar que el usuario sea un paciente
        if not hasattr(self.request.user, 'paciente_perfil'):
            raise PermissionDenied("Acceso no autorizado")
        
        paciente = self.request.user.paciente_perfil
        context['paciente'] = paciente
        
        # Datos para el formulario (mismo formato que la vista mejorada)
        context['familiares'] = [
            ('PADRE', 'Padre'),
            ('MADRE', 'Madre'),
            ('ABUELO_PATERNO', 'Abuelo Paterno'),
            ('ABUELA_PATERNA', 'Abuela Paterna'),
            ('ABUELO_MATERNO', 'Abuelo Materno'),
            ('ABUELA_MATERNA', 'Abuela Materna'),
        ]
        
        context['enfermedades_familiares'] = [
            ('DIABETES', 'Diabetes'),
            ('HIPERTENSION', 'Hipertensión'),
            ('CARDIOPATIAS', 'Cardiopatías'),
            ('ALCOHOLISMO', 'Alcoholismo'),
            ('TABAQUISMO', 'Tabaquismo'),
            ('CANCER', 'Cáncer'),
            ('NEUROLOGICAS', 'Neurológicas'),
            ('HEMATOLOGICAS', 'Hematológicas'),
            ('RENALES', 'Renales'),
            ('VENEREAS', 'Venéreas'),
            ('SOBREPESO', 'Sobrepeso'),
        ]
        
        context['habitos_orales'] = [
            ('SUCCION_DEDO', 'Succión de dedo'),
            ('USO_CHUPON', 'Uso de chupón'),
            ('RESPIRADOR_BUCAL', 'Respirador bucal'),
            ('INTERPOSICION_LINGUAL', 'Interposición lingual'),
            ('DEFICIENCIA_CEPILLADO', 'Deficiencia en el cepillado'),
            ('MORDER_OBJETOS', 'Muerde objetos o uñas'),
        ]
        
        # Cuestionario tradicional
        preguntas = models.PreguntaHistorial.objects.filter(activa=True)
        for pregunta in preguntas:
            models.RespuestaHistorial.objects.get_or_create(
                paciente=paciente, 
                pregunta=pregunta, 
                defaults={'respuesta': ''}
            )
            
        if self.request.POST:
            context['formset'] = forms.RespuestaHistorialFormSet(
                self.request.POST, 
                queryset=paciente.respuestas_historial.all()
            )
        else:
            context['formset'] = forms.RespuestaHistorialFormSet(
                queryset=paciente.respuestas_historial.all()
            )
        
        return context
    
    def post(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        paciente = context['paciente']
        formset = context['formset']
        
        # Procesar datos del formulario mejorado (mismo código que la vista original)
        try:
            # Guardar escalas de dolor
            dolor_numerico = request.POST.get('dolor_numerico')
            dolor_caras = request.POST.get('dolor_caras')
            
            # Guardar antecedentes familiares
            antecedentes_data = {}
            for parentesco, _ in context['familiares']:
                estado = request.POST.get(f'estado_{parentesco}')
                enfermedades = request.POST.getlist(f'enfermedad_{parentesco}')
                observaciones = request.POST.get(f'observaciones_{parentesco}', '')
                
                if estado:
                    antecedentes_data[parentesco] = {
                        'estado': estado,
                        'enfermedades': enfermedades,
                        'observaciones': observaciones
                    }
            
            # Guardar hábitos orales
            habitos_data = {}
            for habito, _ in context['habitos_orales']:
                frecuencia = request.POST.get(f'habito_{habito}')
                if frecuencia and frecuencia != 'NUNCA':
                    habitos_data[habito] = frecuencia
            
            # Guardar signos vitales
            signos_vitales = {
                'estatura': request.POST.get('estatura', ''),
                'peso': request.POST.get('peso', ''),
                'pulso': request.POST.get('pulso', ''),
                'frecuencia_respiratoria': request.POST.get('frecuencia_respiratoria', ''),
                'presion_arterial': request.POST.get('presion_arterial', ''),
                'temperatura': request.POST.get('temperatura', ''),
                'tipo_sangre': request.POST.get('tipo_sangre', ''),
            }
            
            # Crear entrada en el historial clínico
            descripcion_evento = f"""Historial Clínico Auto-Completado por Paciente - {timezone.now().strftime('%d/%m/%Y %H:%M')}
            
    ESCALAS DE DOLOR:
    - Escala numérica: {dolor_numerico or 'No evaluado'}/10
    - Escala de caras Wong Baker: {dolor_caras or 'No evaluado'}
    
    SIGNOS VITALES:
    - Estatura: {signos_vitales.get('estatura', 'No registrado')}
    - Peso: {signos_vitales.get('peso', 'No registrado')}
    - Pulso: {signos_vitales.get('pulso', 'No registrado')} lat/min
    - F.R.: {signos_vitales.get('frecuencia_respiratoria', 'No registrado')} resp/min
    - T.A.: {signos_vitales.get('presion_arterial', 'No registrado')} mmHg
    - Temperatura: {signos_vitales.get('temperatura', 'No registrado')} °C
    - Tipo de sangre: {signos_vitales.get('tipo_sangre', 'No registrado')}
    
    ANTECEDENTES FAMILIARES:
    {self._format_antecedentes_portal(antecedentes_data)}
    
    MALOS HÁBITOS ORALES:
    {self._format_habitos_portal(habitos_data, context['habitos_orales'])}
    
    MOTIVO DE LA CONSULTA:
    {request.POST.get('motivo_consulta', 'No especificado')}
    
    ALERTA MÉDICA:
    {request.POST.get('alerta_medica', 'Ninguna reportada')}
    """
            
            # Crear entrada en historial
            historial = models.HistorialClinico.objects.create(
                paciente=paciente,
                descripcion_evento=descripcion_evento,
                registrado_por=None  # Auto-completado por el paciente
            )
            
            # Guardar formset tradicional
            if formset.is_valid():
                formset.save()
            
            messages.success(request, '¡Tu historial clínico ha sido guardado exitosamente! Nuestro equipo médico lo revisará antes de tu próxima consulta.')
            return redirect('core:portal_historial')
            
        except Exception as e:
            messages.error(request, f'Error al guardar el historial: {str(e)}')
            return self.render_to_response(context)
    
    def _format_antecedentes_portal(self, antecedentes_data):
        """Formatea los antecedentes familiares para el historial del portal"""
        if not antecedentes_data:
            return "No se reportaron antecedentes familiares significativos."
        
        resultado = []
        for parentesco, datos in antecedentes_data.items():
            estado = datos.get('estado', '')
            enfermedades = datos.get('enfermedades', [])
            observaciones = datos.get('observaciones', '')
            
            linea = f"- {parentesco.replace('_', ' ').title()}: {estado}"
            if enfermedades:
                linea += f" | Enfermedades: {', '.join(enfermedades)}"
            if observaciones:
                linea += f" | Obs: {observaciones}"
            resultado.append(linea)
        
        return "\n    ".join(resultado) if resultado else "No se reportaron antecedentes familiares."
    
    def _format_habitos_portal(self, habitos_data, habitos_opciones):
        """Formatea los hábitos orales para el historial del portal"""
        if not habitos_data:
            return "No se reportaron hábitos orales problemáticos."
        
        # Crear diccionario para mapear códigos a nombres
        habitos_nombres = dict(habitos_opciones)
        
        resultado = []
        for habito, frecuencia in habitos_data.items():
            nombre = habitos_nombres.get(habito, habito)
            resultado.append(f"- {nombre}: {frecuencia.replace('_', ' ').title()}")
        
        return "\n    ".join(resultado) if resultado else "No se reportaron hábitos problemáticos."

class ResetPasswordView(TenantLoginRequiredMixin, TemplateView):
    template_name = 'core/usuario_reset_password.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario_a_resetear'] = get_object_or_404(User, pk=self.kwargs['pk'])
        return context

    def post(self, request, *args, **kwargs):
        usuario = get_object_or_404(User, pk=self.kwargs['pk'])
        
        if not request.user.groups.filter(name='Administrador').exists() and not request.user.is_superuser:
            messages.error(request, 'No tienes permiso para realizar esta acción.')
            return redirect('core:usuario_list')

        new_password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
        usuario.set_password(new_password)
        usuario.save()

        messages.success(request, f'La contraseña para {usuario.username} ha sido restablecida a: {new_password}')
        return redirect('core:usuario_edit', pk=usuario.pk)

class CuestionarioHistorialView(TenantLoginRequiredMixin, TemplateView):
    template_name = 'core/cuestionario_historial.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        paciente = get_object_or_404(models.Paciente, pk=self.kwargs['pk'])
        context['paciente'] = paciente
        
        preguntas = models.PreguntaHistorial.objects.filter(activa=True)
        for pregunta in preguntas:
            models.RespuestaHistorial.objects.get_or_create(paciente=paciente, pregunta=pregunta, defaults={'respuesta': ''})
            
        if self.request.POST:
            context['formset'] = forms.RespuestaHistorialFormSet(self.request.POST, queryset=paciente.respuestas_historial.all())
        else:
            context['formset'] = forms.RespuestaHistorialFormSet(queryset=paciente.respuestas_historial.all())
        return context

    def post(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        formset = context['formset']
        
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Cuestionario de historial clínico guardado con éxito.')
            return redirect('core:paciente_detail', pk=self.kwargs['pk'])
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
            return self.render_to_response(context)

class CuestionarioHistorialMejoradoView(TenantLoginRequiredMixin, TemplateView):
    """Vista mejorada del historial clínico con escalas de dolor, antecedentes familiares y hábitos orales"""
    template_name = 'core/cuestionario_historial_mejorado.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        paciente = get_object_or_404(models.Paciente, pk=self.kwargs['pk'])
        context['paciente'] = paciente
        
        # Datos para antecedentes familiares
        context['familiares'] = [
            ('PADRE', 'Padre'),
            ('MADRE', 'Madre'),
            ('ABUELO_PATERNO', 'Abuelo Paterno'),
            ('ABUELA_PATERNA', 'Abuela Paterna'),
            ('ABUELO_MATERNO', 'Abuelo Materno'),
            ('ABUELA_MATERNA', 'Abuela Materna'),
        ]
        
        context['enfermedades_familiares'] = [
            ('DIABETES', 'Diabetes'),
            ('HIPERTENSION', 'Hipertensión'),
            ('CARDIOPATIAS', 'Cardiopatías'),
            ('ALCOHOLISMO', 'Alcoholismo'),
            ('TABAQUISMO', 'Tabaquismo'),
            ('CANCER', 'Cáncer'),
            ('NEUROLOGICAS', 'Neurológicas'),
            ('HEMATOLOGICAS', 'Hematológicas'),
            ('RENALES', 'Renales'),
            ('VENEREAS', 'Venéreas'),
            ('SOBREPESO', 'Sobrepeso'),
        ]
        
        # Datos para hábitos orales
        context['habitos_orales'] = [
            ('SUCCION_DEDO', 'Succión de dedo'),
            ('USO_CHUPON', 'Uso de chupón'),
            ('RESPIRADOR_BUCAL', 'Respirador bucal'),
            ('INTERPOSICION_LINGUAL', 'Interposición lingual'),
            ('DEFICIENCIA_CEPILLADO', 'Deficiencia en el cepillado'),
            ('MORDER_OBJETOS', 'Muerde objetos o uñas'),
        ]
        
        # Cuestionario tradicional
        preguntas = models.PreguntaHistorial.objects.filter(activa=True)
        for pregunta in preguntas:
            models.RespuestaHistorial.objects.get_or_create(paciente=paciente, pregunta=pregunta, defaults={'respuesta': ''})
            
        if self.request.POST:
            context['formset'] = forms.RespuestaHistorialFormSet(self.request.POST, queryset=paciente.respuestas_historial.all())
        else:
            context['formset'] = forms.RespuestaHistorialFormSet(queryset=paciente.respuestas_historial.all())
        
        return context

    def post(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        paciente = context['paciente']
        formset = context['formset']
        
        # Procesar datos del formulario mejorado
        try:
            # Guardar escalas de dolor
            dolor_numerico = request.POST.get('dolor_numerico')
            dolor_caras = request.POST.get('dolor_caras')
            
            # Guardar antecedentes familiares
            antecedentes_data = {}
            for parentesco, _ in context['familiares']:
                estado = request.POST.get(f'estado_{parentesco}')
                enfermedades = request.POST.getlist(f'enfermedad_{parentesco}')
                observaciones = request.POST.get(f'observaciones_{parentesco}', '')
                
                if estado:
                    antecedentes_data[parentesco] = {
                        'estado': estado,
                        'enfermedades': enfermedades,
                        'observaciones': observaciones
                    }
            
            # Guardar hábitos orales
            habitos_data = {}
            for habito, _ in context['habitos_orales']:
                frecuencia = request.POST.get(f'habito_{habito}')
                if frecuencia and frecuencia != 'NUNCA':
                    habitos_data[habito] = frecuencia
            
            # Guardar signos vitales
            signos_vitales = {
                'estatura': request.POST.get('estatura', ''),
                'peso': request.POST.get('peso', ''),
                'pulso': request.POST.get('pulso', ''),
                'frecuencia_respiratoria': request.POST.get('frecuencia_respiratoria', ''),
                'presion_arterial': request.POST.get('presion_arterial', ''),
                'temperatura': request.POST.get('temperatura', ''),
                'tipo_sangre': request.POST.get('tipo_sangre', ''),
            }
            
            # Crear entrada en el historial clínico
            descripcion_evento = f"""Historial Clínico de Atención Inmediata - {timezone.now().strftime('%d/%m/%Y %H:%M')}
            
    ESCALAS DE DOLOR:
    - Escala numérica: {dolor_numerico or 'No evaluado'}/10
    - Escala de caras Wong Baker: {dolor_caras or 'No evaluado'}
    
    SIGNOS VITALES:
    - Estatura: {signos_vitales.get('estatura', 'No registrado')}
    - Peso: {signos_vitales.get('peso', 'No registrado')}
    - Pulso: {signos_vitales.get('pulso', 'No registrado')} lat/min
    - F.R.: {signos_vitales.get('frecuencia_respiratoria', 'No registrado')} resp/min
    - T.A.: {signos_vitales.get('presion_arterial', 'No registrado')} mmHg
    - Temperatura: {signos_vitales.get('temperatura', 'No registrado')} °C
    - Tipo de sangre: {signos_vitales.get('tipo_sangre', 'No registrado')}
    
    ANTECEDENTES FAMILIARES:
    {self._format_antecedentes(antecedentes_data)}
    
    MALOS HÁBITOS ORALES:
    {self._format_habitos(habitos_data, context['habitos_orales'])}
    
    MOTIVO DE LA CONSULTA:
    {request.POST.get('motivo_consulta', 'No especificado')}
    
    ALERTA MÉDICA:
    {request.POST.get('alerta_medica', 'Ninguna reportada')}
    """
            
            # Crear entrada en historial
            historial = models.HistorialClinico.objects.create(
                paciente=paciente,
                descripcion_evento=descripcion_evento,
                registrado_por=getattr(request.user, 'perfildentista', None)
            )
            
            # Guardar formset tradicional
            if formset.is_valid():
                formset.save()
            
            messages.success(request, 'Historial clínico mejorado guardado con éxito.')
            return redirect('core:paciente_detail', pk=paciente.pk)
            
        except Exception as e:
            messages.error(request, f'Error al guardar el historial: {str(e)}')
            return self.render_to_response(context)
    
    def _format_antecedentes(self, antecedentes_data):
        """Formatea los antecedentes familiares para el historial"""
        if not antecedentes_data:
            return "No se reportaron antecedentes familiares significativos."
        
        resultado = []
        for parentesco, datos in antecedentes_data.items():
            estado = datos.get('estado', '')
            enfermedades = datos.get('enfermedades', [])
            observaciones = datos.get('observaciones', '')
            
            linea = f"- {parentesco.replace('_', ' ').title()}: {estado}"
            if enfermedades:
                linea += f" | Enfermedades: {', '.join(enfermedades)}"
            if observaciones:
                linea += f" | Obs: {observaciones}"
            resultado.append(linea)
        
        return "\n    ".join(resultado) if resultado else "No se reportaron antecedentes familiares."
    
    def _format_habitos(self, habitos_data, habitos_opciones):
        """Formatea los hábitos orales para el historial"""
        if not habitos_data:
            return "No se reportaron hábitos orales problemáticos."
        
        # Crear diccionario para mapear códigos a nombres
        habitos_nombres = dict(habitos_opciones)
        
        resultado = []
        for habito, frecuencia in habitos_data.items():
            nombre = habitos_nombres.get(habito, habito)
            resultado.append(f"- {nombre}: {frecuencia.replace('_', ' ').title()}")
        
        return "\n    ".join(resultado) if resultado else "No se reportaron hábitos problemáticos."

class PreguntaHistorialListView(TenantLoginRequiredMixin, ListView):
    model = models.PreguntaHistorial
    template_name = 'core/configuracion/pregunta_list.html'
    context_object_name = 'preguntas'

class PreguntaHistorialCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.PreguntaHistorial
    fields = ['texto', 'tipo', 'opciones', 'orden', 'activa']
    template_name = 'core/configuracion/pregunta_form.html'
    success_url = reverse_lazy('core:pregunta_list')
    success_message = "Pregunta creada con éxito."

class PreguntaHistorialUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.PreguntaHistorial
    fields = ['texto', 'tipo', 'opciones', 'orden', 'activa']
    template_name = 'core/configuracion/pregunta_form.html'
    success_url = reverse_lazy('core:pregunta_list')
    success_message = "Pregunta actualizada con éxito."

class PreguntaHistorialDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = models.PreguntaHistorial
    template_name = 'core/configuracion/pregunta_confirm_delete.html'
    success_url = reverse_lazy('core:pregunta_list')
    context_object_name = 'pregunta'
    success_message = "Pregunta eliminada con éxito."

class ReporteServiciosMasVendidosView(TenantLoginRequiredMixin, ListView):
    model = models.Servicio
    template_name = 'core/reportes/reporte_servicios_vendidos.html'
    context_object_name = 'servicios'

    def get_queryset(self):
        # Corrigiendo la consulta para evitar errores en las relaciones
        return models.Servicio.objects.annotate(
            cantidad_vendida=Count('citas_realizadas', distinct=True),
            ingresos_generados=Sum('precio') * Count('citas_realizadas')
        ).filter(cantidad_vendida__gt=0).order_by('-cantidad_vendida')

class ReporteIngresosPorDentistaView(TenantLoginRequiredMixin, ListView):
    model = models.Pago
    template_name = 'core/reportes/reporte_ingresos_dentista.html'
    context_object_name = 'pagos'

    def get_queryset(self):
        queryset = models.Pago.objects.select_related('paciente', 'cita__dentista').order_by('-fecha_pago')
        form = forms.ReporteIngresosDentistaForm(self.request.GET or None)

        if form and form.is_valid():
            dentista = form.cleaned_data.get('dentista')
            if dentista:
                queryset = queryset.filter(cita__dentista=dentista)
            
            fecha_inicio = form.cleaned_data.get('fecha_inicio')
            if fecha_inicio:
                queryset = queryset.filter(fecha_pago__date__gte=fecha_inicio)

            fecha_fin = form.cleaned_data.get('fecha_fin')
            if fecha_fin:
                queryset = queryset.filter(fecha_pago__date__lte=fecha_fin)
        else:
            # Defaults si no hay parámetros válidos
            from datetime import timedelta
            today = timezone.now().date()
            default_inicio = today - timedelta(days=30)
            default_fin = today
            queryset = queryset.filter(fecha_pago__date__gte=default_inicio, fecha_pago__date__lte=default_fin)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from datetime import timedelta
        today = timezone.now().date()
        default_inicio = today - timedelta(days=30)
        default_fin = today
        # Mantener valores por defecto visibles cuando no hay GET (o solo paginación)
        if 'fecha_inicio' in self.request.GET or 'fecha_fin' in self.request.GET:
            form = forms.ReporteIngresosDentistaForm(self.request.GET)
        else:
            form = forms.ReporteIngresosDentistaForm(initial={'fecha_inicio': default_inicio, 'fecha_fin': default_fin})
        context['form'] = form
        context['default_fecha_inicio'] = default_inicio.isoformat()
        context['default_fecha_fin'] = default_fin.isoformat()
        context['total_ingresos'] = self.get_queryset().aggregate(total=Sum('monto'))['total'] or 0
        return context

class ReporteServiciosVendidosPeriodoView(TenantLoginRequiredMixin, TemplateView):
    template_name = 'core/reportes/reporte_servicios_vendidos_periodo.html'

    def get_context_data(self, **kwargs):
        from datetime import date
        context = super().get_context_data(**kwargs)
        form = forms.ReporteServiciosForm(self.request.GET or None)

        # Valores por defecto si no vienen parámetros
        today = timezone.now().date()
        periodo = 'mes'
        fecha_inicio = None
        fecha_fin = today
        dentista = None

        if form.is_valid():
            periodo = form.cleaned_data.get('periodo') or 'semana'
            fecha_inicio = form.cleaned_data.get('fecha_inicio')
            fecha_fin = form.cleaned_data.get('fecha_fin') or today
            dentista = form.cleaned_data.get('dentista')
        else:
            periodo = (self.request.GET.get('periodo') or 'semana')

        # Default ranges si no se especifican
        if not fecha_inicio:
            if periodo == 'mes':
                # últimos 6 meses
                start_month = (today.replace(day=1) - timedelta(days=150)).replace(day=1)
                fecha_inicio = start_month
            else:
                # últimas 8 semanas
                fecha_inicio = today - timedelta(weeks=8)

        qs = models.Cita.objects.filter(
            fecha_hora__date__gte=fecha_inicio,
            fecha_hora__date__lte=fecha_fin,
            estado__in=['ATN', 'COM'],
            servicios_realizados__isnull=False
        )
        if dentista:
            qs = qs.filter(dentista=dentista)

        resultados = []
        if periodo == 'mes':
            agg = qs.annotate(
                periodo_mes=TruncMonth('fecha_hora'),
                servicio_id=F('servicios_realizados__id'),
                servicio_nombre=F('servicios_realizados__nombre'),
            ).values('periodo_mes', 'servicio_id', 'servicio_nombre').annotate(
                cantidad=Count('id')
            ).order_by('periodo_mes', '-cantidad')

            # Agrupar por mes y tomar top 10
            from collections import defaultdict
            grouped = defaultdict(list)
            for row in agg:
                key = row['periodo_mes'].strftime('%Y-%m') if row['periodo_mes'] else 'N/A'
                grouped[key].append({'servicio': row['servicio_nombre'], 'cantidad': row['cantidad']})
            resultados = [{'periodo': k, 'items': v[:10]} for k, v in sorted(grouped.items())]
        else:
            agg = qs.annotate(
                anio=ExtractYear('fecha_hora'),
                semana=ExtractWeek('fecha_hora'),
                servicio_id=F('servicios_realizados__id'),
                servicio_nombre=F('servicios_realizados__nombre'),
            ).values('anio', 'semana', 'servicio_id', 'servicio_nombre').annotate(
                cantidad=Count('id')
            ).order_by('anio', 'semana', '-cantidad')

            from collections import defaultdict
            grouped = defaultdict(list)
            for row in agg:
                key = f"{row['anio']}-W{int(row['semana']):02d}"
                grouped[key].append({'servicio': row['servicio_nombre'], 'cantidad': row['cantidad']})
            resultados = [{'periodo': k, 'items': v[:10]} for k, v in sorted(grouped.items())]

        context.update({
            'form': form,
            'periodo': periodo,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'resultados': resultados,
        })
        return context

class ReporteIngresosDentistaPeriodoView(TenantLoginRequiredMixin, TemplateView):
    template_name = 'core/reportes/reporte_ingresos_dentista_periodo.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = forms.ReporteIngresosDentistaForm(self.request.GET or None)

        today = timezone.now().date()
        periodo = 'mes'
        fecha_inicio = None
        fecha_fin = today
        dentista = None

        if form.is_valid():
            periodo = form.cleaned_data.get('periodo') or 'semana'
            fecha_inicio = form.cleaned_data.get('fecha_inicio')
            fecha_fin = form.cleaned_data.get('fecha_fin') or today
            dentista = form.cleaned_data.get('dentista')
        else:
            periodo = (self.request.GET.get('periodo') or 'semana')

        if not fecha_inicio:
            if periodo == 'mes':
                start_month = (today.replace(day=1) - timedelta(days=150)).replace(day=1)
                fecha_inicio = start_month
            else:
                fecha_inicio = today - timedelta(weeks=8)

        qs = models.Pago.objects.filter(
            fecha_pago__date__gte=fecha_inicio,
            fecha_pago__date__lte=fecha_fin,
            cita__dentista__isnull=False
        )
        if dentista:
            qs = qs.filter(cita__dentista=dentista)

        resultados = []
        if periodo == 'mes':
            agg = qs.annotate(
                periodo_mes=TruncMonth('fecha_pago'),
                dentista_id=F('cita__dentista__id'),
                dentista_nombre=F('cita__dentista__nombre'),
                dentista_apellido=F('cita__dentista__apellido'),
            ).values('periodo_mes', 'dentista_id', 'dentista_nombre', 'dentista_apellido').annotate(
                total=Sum('monto')
            ).order_by('periodo_mes', '-total')

            from collections import defaultdict
            grouped = defaultdict(list)
            for row in agg:
                key = row['periodo_mes'].strftime('%Y-%m') if row['periodo_mes'] else 'N/A'
                nombre = f"Dr. {row['dentista_nombre']} {row['dentista_apellido']}"
                grouped[key].append({'dentista': nombre, 'total': row['total']})
            resultados = [{'periodo': k, 'items': v} for k, v in sorted(grouped.items())]
        else:
            agg = qs.annotate(
                anio=ExtractYear('fecha_pago'),
                semana=ExtractWeek('fecha_pago'),
                dentista_id=F('cita__dentista__id'),
                dentista_nombre=F('cita__dentista__nombre'),
                dentista_apellido=F('cita__dentista__apellido'),
            ).values('anio', 'semana', 'dentista_id', 'dentista_nombre', 'dentista_apellido').annotate(
                total=Sum('monto')
            ).order_by('anio', 'semana', '-total')

            from collections import defaultdict
            grouped = defaultdict(list)
            for row in agg:
                key = f"{row['anio']}-W{int(row['semana']):02d}"
                nombre = f"Dr. {row['dentista_nombre']} {row['dentista_apellido']}"
                grouped[key].append({'dentista': nombre, 'total': row['total']})
            resultados = [{'periodo': k, 'items': v} for k, v in sorted(grouped.items())]

        context.update({
            'form': form,
            'periodo': periodo,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'resultados': resultados,
        })
        return context
# REEMPLAZA la clase ReporteIngresosView existente por:
class ReporteIngresosView(TenantLoginRequiredMixin, ListView):
    model = models.Pago
    template_name = 'core/reportes/reporte_ingresos.html'
    context_object_name = 'pagos'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = models.Pago.objects.select_related('paciente', 'cita__paciente').order_by('-fecha_pago')
        form = forms.ReporteIngresosForm(self.request.GET or None)
        from datetime import date, timedelta
        today = timezone.now().date()
        default_inicio = today - timedelta(days=30)
        default_fin = today
        if self.request.GET:
            if form and form.is_valid():
                fecha_inicio = form.cleaned_data.get('fecha_inicio') or default_inicio
                fecha_fin = form.cleaned_data.get('fecha_fin') or default_fin
            else:
                fecha_inicio, fecha_fin = default_inicio, default_fin
        else:
            fecha_inicio, fecha_fin = default_inicio, default_fin
        queryset = queryset.filter(fecha_pago__date__gte=fecha_inicio, fecha_pago__date__lte=fecha_fin)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from datetime import timedelta
        today = timezone.now().date()
        default_inicio = today - timedelta(days=30)
        default_fin = today
        # Si el GET no trae fechas (ej. solo paginación), usar defaults visibles
        if 'fecha_inicio' in self.request.GET or 'fecha_fin' in self.request.GET:
            form = forms.ReporteIngresosForm(self.request.GET)
        else:
            form = forms.ReporteIngresosForm(initial={'fecha_inicio': default_inicio, 'fecha_fin': default_fin})
        context['form'] = form
        context['default_fecha_inicio'] = default_inicio.isoformat()
        context['default_fecha_fin'] = default_fin.isoformat()
        context['total_ingresos'] = self.get_queryset().aggregate(total=Sum('monto'))['total'] or 0
        return context
# ==================== VISTAS FALTANTES ====================

# --- PROVEEDORES ---
class ProveedorListView(TenantLoginRequiredMixin, ListView):
    model = models.Proveedor
    template_name = 'core/proveedor_list.html'
    context_object_name = 'proveedores'
    paginate_by = 15

class ProveedorCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, CreateView):
    model = models.Proveedor
    template_name = 'core/proveedor_form.html'
    fields = ['nombre', 'rfc', 'nombre_contacto', 'telefono', 'email', 'direccion_fiscal']
    success_url = reverse_lazy('core:proveedor_list')

class ProveedorUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.Proveedor
    template_name = 'core/proveedor_form.html'
    fields = ['nombre', 'contacto', 'telefono', 'email', 'direccion']
    success_url = reverse_lazy('core:proveedor_list')
    success_message = "Proveedor '%(nombre)s' actualizado con éxito."

class ProveedorDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, DeleteView):
    model = models.Proveedor
    template_name = 'core/proveedor_confirm_delete.html'
    success_url = reverse_lazy('core:proveedor_list')
    context_object_name = 'proveedor'
    
    def form_valid(self, form):
        messages.success(self.request, f"Proveedor '{self.object.nombre}' eliminado con éxito.")
        return super().form_valid(form)

# --- INSUMOS ---
class InsumoListView(TenantLoginRequiredMixin, ListView):
    model = models.Insumo
    template_name = 'core/insumo_list.html'
    context_object_name = 'insumos'
    paginate_by = 15
    
    def get_queryset(self):
        queryset = models.Insumo.objects.select_related('proveedor').prefetch_related(
            'lotes', 'lotes__unidad_dental'
        ).order_by('nombre')
        
        # Filtros
        busqueda = self.request.GET.get('busqueda')
        proveedor_id = self.request.GET.get('proveedor')
        estado_stock = self.request.GET.get('estado_stock')
        
        if busqueda:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(nombre__icontains=busqueda) |
                Q(descripcion__icontains=busqueda)
            )
        
        if proveedor_id:
            queryset = queryset.filter(proveedor_id=proveedor_id)
        
        if estado_stock:
            from django.db.models import F
            if estado_stock == 'critico':
                queryset = queryset.filter(stock=0)
            elif estado_stock == 'bajo':
                queryset = queryset.filter(stock__lte=F('stock_minimo'), stock__gt=0)
            elif estado_stock == 'normal':
                queryset = queryset.filter(stock__gt=F('stock_minimo'))
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Imports necesarios
        from django.db.models import F, Sum
        from datetime import date
        
        # Actualizar stock de todos los insumos
        for insumo in context['insumos']:
            insumo.actualizar_stock_total()
        
        # Estadísticas
        total_insumos = models.Insumo.objects.count()
        stock_critico_count = models.Insumo.objects.filter(stock=0).count()
        stock_bajo_count = models.Insumo.objects.filter(
            stock__lte=F('stock_minimo'), stock__gt=0
        ).count()
        
        # Valor total del inventario
        valor_total = models.Insumo.objects.aggregate(
            valor_total=Sum(F('stock') * F('precio_unitario'))
        )['valor_total'] or 0
        
        # Lista de proveedores para filtros
        proveedores = models.Proveedor.objects.all().order_by('nombre')
        
        context.update({
            'stock_critico_count': stock_critico_count,
            'stock_bajo_count': stock_bajo_count,
            'valor_total_inventario': valor_total,
            'proveedores': proveedores,
            'fecha_actual': date.today(),
        })
        
        return context

class InsumoCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.Insumo
    template_name = 'core/insumo_form.html'
    fields = ['nombre', 'descripcion', 'proveedor', 'stock_minimo', 'requiere_lote_caducidad', 'registro_sanitario']
    success_url = reverse_lazy('core:insumo_list')
    success_message = "Insumo '%(nombre)s' creado con éxito."

class InsumoUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.Insumo
    template_name = 'core/insumo_form.html'
    fields = ['nombre', 'descripcion', 'unidad_medida', 'stock_minimo', 'precio_unitario']
    success_url = reverse_lazy('core:insumo_list')
    success_message = "Insumo '%(nombre)s' actualizado con éxito."

class InsumoDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, DeleteView):
    model = models.Insumo
    template_name = 'core/insumo_confirm_delete.html'
    success_url = reverse_lazy('core:insumo_list')
    context_object_name = 'insumo'
    
    def form_valid(self, form):
        messages.success(self.request, f"Insumo '{self.object.nombre}' eliminado con éxito.")
        return super().form_valid(form)

# --- PAGOS ---
class RegistrarPagoView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.Pago
    form_class = forms.PagoForm
    template_name = 'core/registrar_abono.html'
    success_url = reverse_lazy('core:pago_list')
    success_message = "Pago registrado con éxito."
    
    def get_initial(self):
        initial = super().get_initial()
        if 'paciente_id' in self.kwargs:
            paciente = get_object_or_404(models.Paciente, pk=self.kwargs['paciente_id'])
            initial['cliente'] = paciente
        elif 'cita_id' in self.kwargs:
            cita = get_object_or_404(models.Cita, pk=self.kwargs['cita_id'])
            initial['cita'] = cita
        return initial

class ProcesarPagoView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, CreateView):
    model = models.Pago
    template_name = 'core/pago_abono_paciente.html'
    form_class = forms.PagoForm
    success_url = reverse_lazy('core:citas_pendientes_pago')

    def dispatch(self, request, *args, **kwargs):
        cita_id = self.kwargs.get('pk') or self.request.GET.get('cita')
        if not cita_id:
            messages.error(self.request, "No se especificó una cita válida.")
            return redirect('core:citas_pendientes_pago')
        try:
            models.Cita.objects.get(id=cita_id)
        except models.Cita.DoesNotExist:
            messages.error(self.request, f"No se encontró una cita con ID {cita_id}. Por favor, seleccione una cita válida.")
            return redirect('core:citas_pendientes_pago')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        cita_id = self.kwargs.get('pk') or self.request.GET.get('cita')
        cita = models.Cita.objects.get(id=cita_id)
        # Pasar el paciente al formulario para ocultar el selector y ajustar límites
        kwargs['paciente'] = cita.paciente
        kwargs['permitir_factura'] = True
        # En pagos por cita, no permitir seleccionar destino (ya es esa cita)
        kwargs['permitir_destino'] = False
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cita_id = self.kwargs.get('pk') or self.request.GET.get('cita')
        context['cita'] = models.Cita.objects.get(id=cita_id)  # Ya validado en dispatch
        context['paciente'] = context['cita'].paciente
        context['datos_fiscales_form'] = forms.DatosFiscalesForm()
        context['has_datos_fiscales'] = models.DatosFiscales.objects.filter(paciente=context['paciente']).exists()
        return context

    def form_valid(self, form):
        cita_id = self.kwargs.get('pk') or self.request.GET.get('cita')
        try:
            pago = form.save(commit=False)
            pago.cita = models.Cita.objects.get(id=cita_id)
            pago.paciente = pago.cita.paciente
            pago.save()
        except models.Cita.DoesNotExist:
            messages.error(self.request, f"No se encontró una cita con ID {cita_id}. Por favor, seleccione una cita válida.")
            return self.form_invalid(form)

        # Verificar si desea facturar
        if form.cleaned_data['desea_factura']:
            # Marcar la cita como requerida para facturación
            try:
                pago.cita.requiere_factura = True
                pago.cita.save(update_fields=['requiere_factura'])
            except Exception:
                pass
            try:
                datos_fiscales = models.DatosFiscales.objects.get(paciente=pago.paciente)
                if not all([datos_fiscales.rfc, datos_fiscales.razon_social, (datos_fiscales.direccion_completa or datos_fiscales.domicilio_fiscal)]):
                    messages.error(self.request, "El paciente no tiene datos fiscales completos. Por favor, complételos.")
                    return self.form_invalid(form)
                # Aquí iría la lógica para generar la factura
                messages.success(self.request, "Pago registrado y factura solicitada con éxito.")
            except models.DatosFiscales.DoesNotExist:
                datos_fiscales_form = forms.DatosFiscalesForm(self.request.POST)
                if datos_fiscales_form.is_valid():
                    datos_fiscales = datos_fiscales_form.save(commit=False)
                    datos_fiscales.paciente = pago.paciente
                    datos_fiscales.save()
                    messages.success(self.request, "Pago registrado y datos fiscales guardados. Factura solicitada.")
                else:
                    messages.error(self.request, "Por favor, complete los datos fiscales.")
                    return self.render_to_response(
                        self.get_context_data(form=form, datos_fiscales_form=datos_fiscales_form)
                    )
        else:
            messages.success(self.request, "Pago registrado con éxito.")

        # Actualizar saldo global
        pago.paciente.actualizar_saldo_global()
        return super().form_valid(form)

# --- REPORTES ---
class ReporteSaldosView(TenantLoginRequiredMixin, ListView):
    model = models.Paciente
    template_name = 'core/reportes/reporte_saldos.html'
    context_object_name = 'pacientes'

    def get_queryset(self):
        pacientes = models.Paciente.objects.all()
        for paciente in pacientes:
            paciente.actualizar_saldo_global()
        return pacientes.filter(saldo_global__gt=0).select_related('usuario')
    
class ReporteFacturacionView(TenantLoginRequiredMixin, ListView):
    model = models.Cita
    template_name = 'core/reportes/reporte_facturacion.html'
    context_object_name = 'citas_facturacion'
    
    def get_queryset(self):
        # Mostrar citas con factura solicitada (flag) o con pagos que tengan datos SAT (fallback)
        base = models.Cita.objects.select_related('paciente').order_by('-fecha_hora')
        queryset = base.filter(
            Q(requiere_factura=True) | Q(pagos__forma_pago_sat__isnull=False)
        ).distinct().prefetch_related(
            'servicios_realizados',
            'pagos__forma_pago_sat',
            'pagos__metodo_sat'
        )
        
        # Aplicar filtros basados en parámetros GET
        nombre_paciente = self.request.GET.get('nombre_paciente', '').strip()
        if nombre_paciente:
            queryset = queryset.filter(
                Q(paciente__usuario__first_name__icontains=nombre_paciente) |
                Q(paciente__usuario__last_name__icontains=nombre_paciente) |
                Q(paciente__nombre__icontains=nombre_paciente) |
                Q(paciente__apellido__icontains=nombre_paciente)
            )
        
        rfc_filtro = self.request.GET.get('rfc', '').strip()
        if rfc_filtro:
            queryset = queryset.filter(
                paciente__datos_fiscales__rfc__icontains=rfc_filtro
            )
        
        fecha_inicio = self.request.GET.get('fecha_inicio', '').strip()
        if fecha_inicio:
            try:
                from datetime import datetime
                fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                queryset = queryset.filter(fecha_hora__date__gte=fecha_inicio_dt)
            except ValueError:
                pass
        
        fecha_fin = self.request.GET.get('fecha_fin', '').strip()
        if fecha_fin:
            try:
                from datetime import datetime
                fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                queryset = queryset.filter(fecha_hora__date__lte=fecha_fin_dt)
            except ValueError:
                pass
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        citas_qs = list(context.get('citas_facturacion') or context.get('object_list'))
        
        # Enriquecer cada cita con su último pago para evitar problemas en template
        for cita in citas_qs:
            ultimo_pago = cita.pagos.order_by('-fecha_pago').first()
            cita.ultimo_pago = ultimo_pago
        
        # Pasar el nombre esperado por la plantilla actual
        context['citas_a_facturar'] = citas_qs
        # Calcular total a facturar como suma de montos pagados en estas citas
        total = models.Pago.objects.filter(cita__in=citas_qs).aggregate(total=Sum('monto'))['total'] or 0
        context['total_a_facturar'] = total
        return context
        
class CitaDetailView(TenantLoginRequiredMixin, DetailView):
    model = models.Cita
    template_name = 'core/cita_detail.html'
    context_object_name = 'cita'

class CitaManageView(TenantLoginRequiredMixin, DetailView):
    model = models.Cita
    template_name = 'core/cita_manage.html'
    context_object_name = 'cita'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['historial_form'] = forms.HistorialClinicoForm()
        context['puede_facturar'] = self.request.user.groups.filter(name__in=['Administrador','Recepcionista']).exists()
        
        # Agregar diagnósticos para el formulario de tratamientos
        context['diagnosticos'] = models.Diagnostico.objects.all().order_by('nombre')
        
        return context
        
    def post(self, request, *args, **kwargs):
        """Manejar registro de tratamientos y entradas de historial clínico"""
        cita = self.get_object()
        action = request.POST.get('action', 'tratamiento')
        
        if action == 'historial':
            return self._handle_historial_entry(request, cita)
        else:
            return self._handle_tratamiento(request, cita)
    
    def _handle_historial_entry(self, request, cita):
        """Manejar creación de entrada de historial clínico"""
        try:
            # Obtener dentista
            perfil_dentista = models.PerfilDentista.objects.get(usuario=request.user)
        except models.PerfilDentista.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Sin permisos de dentista'}, status=403)
        
        try:
            # Importar función helper
            from .models import crear_entrada_historial_clinico
            
            tipo_registro = request.POST.get('tipo_registro', 'CONSULTA')
            descripcion = request.POST.get('descripcion_evento', '').strip()
            
            if not descripcion:
                return JsonResponse({'success': False, 'error': 'La descripción es obligatoria'}, status=400)
            
            # Crear entrada
            entrada = crear_entrada_historial_clinico(
                paciente=cita.paciente,
                tipo_registro=tipo_registro,
                descripcion=descripcion,
                dentista=perfil_dentista,
                cita=cita
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Entrada de {entrada.get_tipo_registro_display().lower()} agregada correctamente',
                'entrada_id': entrada.id
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    def _handle_tratamiento(self, request, cita):
        """Manejar registro de tratamientos en la cita"""
        # Verificar permisos (dentista de la cita o admin)
        try:
            perfil_dentista = models.PerfilDentista.objects.get(usuario=request.user)
            if cita.dentista != perfil_dentista and not request.user.is_superuser:
                return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)
        except models.PerfilDentista.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)
        
        try:
            # Importar las funciones helper que creamos
            from .models import procesar_tratamiento_cita
            
            # Obtener datos del formulario
            dientes_tratados = request.POST.get('dientes_tratados', '').strip()
            descripcion = request.POST.get('descripcion', '').strip()
            estado_inicial = request.POST.get('estado_inicial', '').strip()
            estado_final = request.POST.get('estado_final', '').strip()
            diagnostico_final_id = request.POST.get('diagnostico_final_id')
            trabajo_pendiente = request.POST.get('trabajo_pendiente', '').strip()
            requiere_seguimiento = request.POST.get('requiere_seguimiento') == 'on'
            fecha_seguimiento = request.POST.get('fecha_seguimiento', '').strip()
            
            # Validaciones
            if not all([dientes_tratados, descripcion, estado_inicial, estado_final, diagnostico_final_id]):
                return JsonResponse({'success': False, 'error': 'Faltan campos obligatorios'}, status=400)
            
            # Validar dientes (importar la función de validación)
            from .models import validar_numero_diente_fdi
            try:
                dientes_list = [int(d.strip()) for d in dientes_tratados.split(',') if d.strip()]
                for diente in dientes_list:
                    if not validar_numero_diente_fdi(diente):
                        return JsonResponse({'success': False, 'error': f'Diente inválido: {diente}'}, status=400)
            except ValueError:
                return JsonResponse({'success': False, 'error': 'Formato de dientes inválido'}, status=400)
            
            # Obtener diagnóstico final
            try:
                diagnostico_final = models.Diagnostico.objects.get(id=diagnostico_final_id)
            except models.Diagnostico.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Diagnóstico no encontrado'}, status=400)
            
            # Procesar fecha de seguimiento
            from datetime import datetime
            fecha_seguimiento_obj = None
            if fecha_seguimiento:
                try:
                    fecha_seguimiento_obj = datetime.strptime(fecha_seguimiento, '%Y-%m-%d').date()
                except ValueError:
                    return JsonResponse({'success': False, 'error': 'Fecha de seguimiento inválida'}, status=400)
            
            # Procesar tratamiento usando nuestra función helper
            with transaction.atomic():
                tratamiento, estados_actualizados, cambios_realizados = procesar_tratamiento_cita(
                    cita=cita,
                    dientes_tratados_str=dientes_tratados,
                    descripcion_tratamiento=descripcion,
                    estado_inicial_desc=estado_inicial,
                    estado_final_desc=estado_final,
                    diagnostico_final=diagnostico_final,
                    servicios_ids=None,  # Se puede agregar más tarde si es necesario
                    trabajo_pendiente=trabajo_pendiente,
                    requiere_seguimiento=requiere_seguimiento,
                    fecha_seguimiento=fecha_seguimiento_obj
                )
            
            # Respuesta exitosa
            response_data = {
                'success': True,
                'message': f'Tratamiento registrado correctamente. {cambios_realizados} dientes actualizados.',
                'tratamiento_id': tratamiento.id,
                'dientes_tratados': len(dientes_list),
                'cambios_realizados': cambios_realizados
            }
            
            return JsonResponse(response_data)
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

class CitaCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.Cita
    form_class = forms.CitaForm
    template_name = 'core/cita_form.html'
    success_message = "Cita creada exitosamente."
    
    def get_success_url(self):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return reverse_lazy('core:agenda')
        return reverse_lazy('core:cita_list')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            cita = form.save()
            return JsonResponse({
                'success': True,
                'message': 'Cita creada exitosamente.',
                'cita_id': cita.id
            })
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # DEBUGGING: Log form errors para diagnosticar problema 400
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"CitaCreateView form_invalid - Errores: {form.errors}")
        logger.error(f"CitaCreateView form_invalid - Datos POST: {dict(self.request.POST)}")
        
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'errors': form.errors
            }, status=400)
        return super().form_invalid(form)

class CitaUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.Cita
    form_class = forms.CitaForm
    template_name = 'core/cita_form.html'
    success_message = "Cita actualizada exitosamente."
    
    def get_success_url(self):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return reverse_lazy('core:agenda')
        return reverse_lazy('core:cita_list')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            cita = form.save()
            return JsonResponse({
                'success': True,
                'message': 'Cita actualizada exitosamente.',
                'cita_id': cita.id
            })
        return super().form_valid(form)
    
    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            # Devolver 200 para que el frontend no lo trate como error de red
            return JsonResponse({
                'success': False,
                'errors': form.errors
            }, status=200)
        return super().form_invalid(form)

class CitaDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = models.Cita
    success_url = reverse_lazy('core:cita_list')
    success_message = "Cita eliminada exitosamente."
    
    def delete(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            self.object = self.get_object()
            self.object.delete()
            return JsonResponse({
                'success': True,
                'message': 'Cita eliminada exitosamente.'
            })
        return super().delete(request, *args, **kwargs)
    
    
class SaldosPendientesListView(TenantLoginRequiredMixin, ListView):
    """
    Vista para mostrar pacientes con saldo global pendiente
    """
    model = models.Paciente
    template_name = 'core/saldos_pendientes.html'
    context_object_name = 'pacientes'
    paginate_by = 20

    def get_queryset(self):
        # Obtener pacientes con saldo global > 0
        queryset = models.Paciente.objects.filter(saldo_global__gt=0).select_related('usuario')
        
        # Filtros
        paciente_nombre = self.request.GET.get('paciente', '')
        if paciente_nombre:
            queryset = queryset.filter(
                models.Q(nombre__icontains=paciente_nombre) | 
                models.Q(apellido__icontains=paciente_nombre)
            )
        
        # Ordenar por saldo mayor primero
        return queryset.order_by('-saldo_global', 'apellido', 'nombre')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Filtros aplicados (para mantener en los links de paginación)
        context['filtros_actuales'] = {
            'paciente': self.request.GET.get('paciente', ''),
        }
        
        # Estadísticas
        context['total_pacientes_con_saldo'] = self.get_queryset().count()
        context['total_saldo_pendiente'] = self.get_queryset().aggregate(
            total=Sum('saldo_global')
        )['total'] or 0
        
        return context

class PacientesPendientesPagoListView(TenantLoginRequiredMixin, ListView):
    """
    Vista alternativa que muestra pacientes con saldo global (mantiene compatibilidad)
    """
    model = models.Paciente
    template_name = 'core/pacientes_pendientes_pago.html'
    context_object_name = 'pacientes'
    paginate_by = 20

    def get_queryset(self):
        # Obtener pacientes con saldo global > 0 y también sus citas ATN
        pacientes_con_saldo = models.Paciente.objects.filter(
            saldo_global__gt=0
        ).select_related('usuario').prefetch_related(
            models.Prefetch(
                'citas',
                queryset=models.Cita.objects.filter(estado='ATN').select_related('dentista', 'unidad_dental')
            )
        )
        
        # Filtros
        paciente_nombre = self.request.GET.get('paciente', '')
        if paciente_nombre:
            pacientes_con_saldo = pacientes_con_saldo.filter(
                models.Q(nombre__icontains=paciente_nombre) | 
                models.Q(apellido__icontains=paciente_nombre)
            )
        
        return pacientes_con_saldo.order_by('-saldo_global', 'apellido', 'nombre')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Solo admin y recepción pueden filtrar por dentista
        context['puede_filtrar_dentista'] = user.groups.filter(
            name__in=['Administrador', 'Recepcionista']
        ).exists()
        
        if context['puede_filtrar_dentista']:
            context['dentistas'] = models.PerfilDentista.objects.filter(activo=True)
        
        # Filtros aplicados (para mantener en los links de paginación)
        context['filtros_actuales'] = {
            'paciente': self.request.GET.get('paciente', ''),
        }
        
        # Estadísticas
        context['total_pacientes_con_saldo'] = self.get_queryset().count()
        context['total_saldo_pendiente'] = self.get_queryset().aggregate(
            total=Sum('saldo_global')
        )['total'] or 0
        
        return context
    
    
class RegistrarPagoPacienteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, CreateView):
    model = models.Pago
    form_class = forms.PagoForm
    template_name = 'core/registrar_pago_paciente.html'
    success_url = reverse_lazy('core:saldos_pendientes')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        paciente_id = self.kwargs.get('paciente_id')
        self.paciente = get_object_or_404(models.Paciente, id=paciente_id)
        # Asegurar saldo actualizado antes de renderizar
        try:
            self.paciente.actualizar_saldo_global()
        except Exception:
            pass
        kwargs['paciente'] = self.paciente
        # En esta pantalla sí permitimos facturar y seleccionar destino del pago
        kwargs['permitir_factura'] = True
        kwargs['permitir_destino'] = True
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['paciente'] = self.paciente
        context['saldo_actual'] = getattr(self.paciente, 'saldo_global', 0)
        return context

    def form_valid(self, form):
        pago = form.save(commit=False)
        pago.paciente = self.paciente
        # Asignar destino si se seleccionó aplicar a una cita específica
        aplicar_a = form.cleaned_data.get('aplicar_a') if 'aplicar_a' in form.cleaned_data else None
        if aplicar_a and isinstance(aplicar_a, str) and aplicar_a.startswith('cita:'):
            try:
                cita_id = int(aplicar_a.split(':', 1)[1])
                pago.cita = models.Cita.objects.get(id=cita_id, paciente=self.paciente)
            except Exception:
                pago.cita = None
        pago.save()
        # Actualizar saldo global del paciente de forma segura
        self.paciente.actualizar_saldo_global()
        # Si se desea facturar y no hay datos fiscales, redirigir para capturarlos
        if form.cleaned_data.get('desea_factura'):
            # Marcar la cita como requerida si aplica
            try:
                if pago.cita_id:
                    pago.cita.requiere_factura = True
                    pago.cita.save(update_fields=['requiere_factura'])
            except Exception:
                pass
            if not models.DatosFiscales.objects.filter(paciente=self.paciente).exists():
                messages.info(self.request, 'Para facturar, primero complete los datos fiscales del paciente.')
                return redirect('core:paciente_datos_fiscales', pk=self.paciente.pk)
        return super().form_valid(form)

@tenant_login_required
def get_servicios_for_dentista_api(request, dentista_id):
    try:
        dentista = models.PerfilDentista.objects.get(pk=dentista_id)
        
        # Obtener los IDs de las especialidades directas del dentista
        especialidades_directas_ids = set(dentista.especialidades.values_list('id', flat=True))
        print(f"Espec. directas: {especialidades_directas_ids}")
        # Obtener los IDs de las especialidades incluidas
        especialidades_incluidas_ids = set()
        for esp in dentista.especialidades.prefetch_related('especialidades_incluidas').all():
            for incluida in esp.especialidades_incluidas.all():
                especialidades_incluidas_ids.add(incluida.id)
        print(f"Espec. incluidas: {especialidades_incluidas_ids}")

        # Combinar todos los IDs de especialidades relevantes
        todas_las_especialidades_ids = especialidades_directas_ids.union(especialidades_incluidas_ids)
        print(f"Todas las espec.: {todas_las_especialidades_ids}")
        # Filtrar los servicios que son activos y pertenecen a cualquiera de esas especialidades
        servicios = models.Servicio.objects.filter(
            especialidad__id__in=todas_las_especialidades_ids,
            activo=True
        ).values('id', 'nombre').distinct()
        print(f"Servicios: {list(servicios)}")
        return JsonResponse(list(servicios), safe=False)
        
    except models.PerfilDentista.DoesNotExist:
        print(f"Dentista {dentista_id} no encontrado")
        return JsonResponse({'error': 'Dentista no encontrado'}, status=404)
    except Exception as e:
        print(f"Error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@tenant_login_required
def paciente_pagos_api(request, paciente_id):
    paciente = get_object_or_404(models.Paciente, id=paciente_id)
    limit = int(request.GET.get('limit', '5') or '5')
    pagos = models.Pago.objects.filter(paciente=paciente).order_by('-fecha_pago')[:limit]
    data = [
        {
            'id': p.id,
            'fecha': p.fecha_pago.strftime('%Y-%m-%d %H:%M'),
            'monto': f"{p.monto:.2f}",
            'metodo': p.metodo_pago,
            'cita_id': p.cita_id,
        }
        for p in pagos
    ]
    return JsonResponse({'pagos': data})

@tenant_login_required
def paciente_saldo_api(request, paciente_id):
    paciente = get_object_or_404(models.Paciente, id=paciente_id)
    try:
        paciente.actualizar_saldo_global()
    except Exception:
        pass
    # Devolver como string para evitar problemas de serialización de Decimal
    return JsonResponse({'saldo': f"{paciente.saldo_global:.2f}"})

@tenant_login_required
def get_horario_dentista_api(request, dentista_id):
    try:
        dentista = models.PerfilDentista.objects.get(pk=dentista_id)
        horarios = models.HorarioLaboral.objects.filter(dentista=dentista, activo=True)
        data = [
            {
                'dia': h.get_dia_semana_display(),
                'hora_inicio': h.hora_inicio.strftime('%H:%M'),
                'hora_fin': h.hora_fin.strftime('%H:%M')
            } for h in horarios
        ]
        return JsonResponse(data, safe=False)
    except models.PerfilDentista.DoesNotExist:
        return JsonResponse({'error': 'Dentista no encontrado'}, status=404)

# ==================== VISTAS FALTANTES IMPORTANTES ====================

# --- REGISTRAR ABONO ---
class RegistrarAbonoView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    """
    Vista para registrar un abono general (no vinculado a una cita específica)
    """
    model = models.Pago
    form_class = forms.PagoForm
    template_name = 'core/registrar_abono.html'
    success_url = reverse_lazy('core:pago_list')
    success_message = "Abono registrado con éxito."
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # En abonos generales no permitimos facturación desde aquí
        kwargs['permitir_factura'] = False
        return kwargs
    
    def form_valid(self, form):
        pago = form.save(commit=False)
        # Un abono no está vinculado a una cita específica
        pago.cita = None
        pago.save()
        
        # Actualizar saldo del paciente si se especifica
        if pago.paciente:
            from . import services
            services.PacienteService.actualizar_saldo_global(pago.paciente)
        
        return super().form_valid(form)

# --- FINALIZAR CITA FORM CONTENT ---
@tenant_login_required
def finalizar_cita_form_content(request, pk):
    """
    Devuelve el contenido del formulario para finalizar cita (AJAX)
    """
    cita = get_object_or_404(models.Cita, pk=pk)
    
    if request.method == 'GET':
        form = forms.FinalizarCitaForm(instance=cita)
        html = render_to_string(
            'core/finalizar_cita_form_content.html',
            {'form': form, 'cita': cita},
            request=request
        )
        return JsonResponse({'success': True, 'html': html})
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

# --- GESTIÓN DE UNIDADES DENTALES ---
class UnidadDentalListView(TenantLoginRequiredMixin, ListView):
    model = models.UnidadDental
    template_name = 'core/unidad_dental_list.html'
    context_object_name = 'unidades'
    paginate_by = 15

class UnidadDentalDetailView(TenantLoginRequiredMixin, DetailView):
    model = models.UnidadDental
    template_name = 'core/unidad_dental_detail.html'
    context_object_name = 'unidad'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        unidad = self.object
        # Stock de insumos en esta unidad
        context['stock_insumos'] = models.LoteInsumo.objects.filter(unidad_dental=unidad).select_related('insumo').order_by('insumo__nombre')
        # Dentistas que han tenido citas en esta unidad
        dentistas_ids = models.Cita.objects.filter(unidad_dental=unidad).values_list('dentista_id', flat=True).distinct()
        context['dentistas_asociados'] = models.PerfilDentista.objects.filter(id__in=dentistas_ids).order_by('apellido', 'nombre')
        return context

class UnidadDentalCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.UnidadDental
    template_name = 'core/unidad_dental_form.html'
    form_class = forms.UnidadDentalForm
    success_url = reverse_lazy('core:unidad_dental_list')
    success_message = "Unidad dental '%(nombre)s' creada con éxito."

class UnidadDentalUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.UnidadDental
    template_name = 'core/unidad_dental_form.html'
    form_class = forms.UnidadDentalForm
    success_url = reverse_lazy('core:unidad_dental_list')
    success_message = "Unidad dental '%(nombre)s' actualizada con éxito."

class UnidadDentalDeleteView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = models.UnidadDental
    template_name = 'core/unidad_dental_confirm_delete.html'
    success_url = reverse_lazy('core:unidad_dental_list')
    context_object_name = 'unidad'
    success_message = "Unidad dental eliminada con éxito."

# --- GESTIÓN DE HORARIOS ---
class GestionarHorarioView(TenantLoginRequiredMixin, TemplateView):
    template_name = 'core/gestionar_horario.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dentista_id = self.kwargs.get('dentista_id')
        dentista = get_object_or_404(models.PerfilDentista, id=dentista_id)
        context['dentista'] = dentista
        
        # Determinar si el usuario puede editar
        user = self.request.user
        context['puede_editar'] = (
            user.groups.filter(name='Administrador').exists() or
            user.is_superuser or
            (hasattr(user, 'perfil_dentista') and user.perfil_dentista == dentista)
        )
        
        # Crear formset para horarios
        if self.request.POST:
            context['formset'] = forms.HorarioLaboralFormSet(
                self.request.POST, 
                queryset=models.HorarioLaboral.objects.filter(dentista=dentista),
                prefix='horarios'
            )
        else:
            context['formset'] = forms.HorarioLaboralFormSet(
                queryset=models.HorarioLaboral.objects.filter(dentista=dentista),
                prefix='horarios'
            )
        
        return context
    
    def post(self, request, *args, **kwargs):
        dentista_id = self.kwargs.get('dentista_id')
        dentista = get_object_or_404(models.PerfilDentista, id=dentista_id)
        
        # Verificar permisos
        user = request.user
        puede_editar = (
            user.groups.filter(name='Administrador').exists() or
            user.is_superuser or
            (hasattr(user, 'perfil_dentista') and user.perfil_dentista == dentista)
        )
        
        if not puede_editar:
            messages.error(request, 'No tienes permisos para editar este horario.')
            from django.http import HttpResponseRedirect
            tenant_prefix = request.session.get('tenant_prefix', '')
            redirect_url = f'/{tenant_prefix}/dentistas/{dentista_id}/horario/' if tenant_prefix else f'/dentistas/{dentista_id}/horario/'
            return HttpResponseRedirect(redirect_url)
        
        # Procesar formset
        formset = forms.HorarioLaboralFormSet(
            request.POST,
            queryset=models.HorarioLaboral.objects.filter(dentista=dentista),
            prefix='horarios'
        )
        
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                instance.dentista = dentista
                instance.save()
            formset.save_m2m()
            
            # Eliminar los marcados para eliminación
            for obj in formset.deleted_objects:
                obj.delete()
            
            messages.success(request, 'Horarios actualizados con éxito.')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
        
        from django.http import HttpResponseRedirect
        tenant_prefix = request.session.get('tenant_prefix', '')
        redirect_url = f'/{tenant_prefix}/dentistas/{dentista_id}/horario/' if tenant_prefix else f'/dentistas/{dentista_id}/horario/'
        return HttpResponseRedirect(redirect_url)

# === VISTAS PARA CUESTIONARIO DE HISTORIAL CLÍNICO ===

class CuestionarioHistorialListView(TenantLoginRequiredMixin, TemplateView):
    template_name = 'core/cuestionario/lista.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener pacientes con sus últimos cuestionarios
        pacientes = models.Paciente.objects.prefetch_related(
            'cuestionarios_completados', 'respuestas_historial'
        ).order_by('apellido', 'nombre')
        
        # Enriquecer cada paciente con información del cuestionario
        for paciente in pacientes:
            ultimo_cuestionario = paciente.cuestionarios_completados.first()
            paciente.ultimo_cuestionario = ultimo_cuestionario
            paciente.tiene_cuestionario = ultimo_cuestionario is not None
            paciente.alertas_count = 0  # Placeholder para contar alertas
            
            if ultimo_cuestionario:
                # Contar respuestas críticas para las alertas
                respuestas_criticas = paciente.respuestas_historial.filter(
                    pregunta__importancia='CRITICA'
                ).count()
                paciente.alertas_count = respuestas_criticas
        
        context['pacientes'] = pacientes
        
        # Estadísticas para el dashboard
        context['total_pacientes'] = pacientes.count()
        context['cuestionarios_completos'] = sum(1 for p in pacientes if p.tiene_cuestionario)
        context['cuestionarios_pendientes'] = context['total_pacientes'] - context['cuestionarios_completos']
        context['cuestionarios_con_alertas'] = sum(1 for p in pacientes if p.ultimo_cuestionario and p.ultimo_cuestionario.tiene_alertas)
        
        return context

class CompletarCuestionarioView(TenantLoginRequiredMixin, TemplateView):
    template_name = 'core/cuestionario/completar.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        paciente_id = self.kwargs['paciente_id']
        context['paciente'] = get_object_or_404(models.Paciente, id=paciente_id)
        
        # Obtener categorías y preguntas organizadas
        categorias = models.CategoriaHistorial.objects.filter(
            activa=True,
            preguntas__activa=True
        ).prefetch_related(
            models.Prefetch(
                'preguntas',
                queryset=models.PreguntaHistorial.objects.filter(activa=True).order_by('orden')
            )
        ).distinct().order_by('orden')
        
        # Agregar preguntas ordenadas para cada categoria
        for categoria in categorias:
            categoria.preguntas_ordenadas = categoria.preguntas.all()
        
        context['categorias'] = categorias
        context['categorias_con_preguntas'] = categorias  # Para compatibilidad con template
        
        if self.request.POST:
            context['form'] = forms.CuestionarioHistorialForm(
                paciente=context['paciente'], 
                data=self.request.POST
            )
        else:
            context['form'] = forms.CuestionarioHistorialForm(paciente=context['paciente'])
        
        return context
    
    def post(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        form = context['form']
        
        if form.is_valid():
            cuestionario_completado, respuestas = form.save(
                completado_por=request.user
            )
            
            messages.success(
                request, 
                f'Cuestionario completado exitosamente. '
                f'Se guardaron {len(respuestas)} respuestas.'
            )
            
            if cuestionario_completado.tiene_alertas:
                messages.warning(
                    request,
                    'Se generaron alertas médicas que requieren atención. '
                    'Revise el historial del paciente.'
                )
            
            return redirect('core:cuestionario_detalle', paciente_id=context['paciente'].id)
        
        messages.error(request, 'Por favor, corrija los errores en el formulario.')
        return self.render_to_response(context)

class CuestionarioDetalleView(TenantLoginRequiredMixin, DetailView):
    model = models.Paciente
    template_name = 'core/cuestionario/detalle.html'
    context_object_name = 'paciente'
    pk_url_kwarg = 'paciente_id'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        paciente = self.object
        
        # Obtener respuestas organizadas por categoría
        respuestas = models.RespuestaHistorial.objects.filter(
            paciente=paciente
        ).select_related('pregunta__categoria').order_by(
            'pregunta__categoria__orden', 'pregunta__orden'
        )
        
        # Organizar respuestas por categoría
        respuestas_por_categoria = {}
        for respuesta in respuestas:
            categoria = respuesta.pregunta.categoria
            if categoria not in respuestas_por_categoria:
                respuestas_por_categoria[categoria] = []
            respuestas_por_categoria[categoria].append(respuesta)
        
        context['respuestas_por_categoria'] = respuestas_por_categoria
        
        # Último cuestionario completado
        context['ultimo_cuestionario'] = paciente.cuestionarios_completados.first()
        
        # Respuestas críticas
        respuestas_criticas = [r for r in respuestas if r.es_respuesta_critica()]
        context['respuestas_criticas'] = respuestas_criticas
        
        # Respuestas que requieren seguimiento
        respuestas_seguimiento = [
            r for r in respuestas 
            if r.pregunta.requiere_seguimiento and r.respuesta.lower() in ['sí', 'si']
        ]
        context['respuestas_seguimiento'] = respuestas_seguimiento
        
        return context

# === VISTAS DE ADMINISTRACIÓN DEL CUESTIONARIO ===

class CategoriaHistorialListView(TenantLoginRequiredMixin, ListView):
    model = models.CategoriaHistorial
    template_name = 'core/cuestionario/admin/categoria_list.html'
    context_object_name = 'categorias'
    
    def get_queryset(self):
        return models.CategoriaHistorial.objects.prefetch_related(
            'preguntas'
        ).order_by('orden', 'nombre')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calcular estadísticas para cada categoría
        total_preguntas_global = 0
        categorias_activas = 0
        
        for categoria in context['categorias']:
            categoria.total_preguntas = categoria.preguntas.count()
            categoria.preguntas_activas = categoria.preguntas.filter(activa=True).count()
            categoria.preguntas_inactivas = categoria.total_preguntas - categoria.preguntas_activas
            
            total_preguntas_global += categoria.total_preguntas
            if categoria.activa:
                categorias_activas += 1
        
        # Agregar estadísticas globales al contexto
        context['total_preguntas_global'] = total_preguntas_global
        context['categorias_activas'] = categorias_activas
            
        return context

class CategoriaHistorialCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.CategoriaHistorial
    form_class = forms.CategoriaHistorialForm
    template_name = 'core/cuestionario/admin/categoria_form.html'
    success_url = reverse_lazy('core:categoria_historial_list')
    success_message = "Categoría '%(nombre)s' creada exitosamente."

class CategoriaHistorialUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.CategoriaHistorial
    form_class = forms.CategoriaHistorialForm
    template_name = 'core/cuestionario/admin/categoria_form.html'
    success_url = reverse_lazy('core:categoria_historial_list')
    success_message = "Categoría '%(nombre)s' actualizada exitosamente."

class PreguntaHistorialListView(TenantLoginRequiredMixin, ListView):
    model = models.PreguntaHistorial
    template_name = 'core/cuestionario/admin/pregunta_list.html'
    context_object_name = 'preguntas'
    
    def get_queryset(self):
        return models.PreguntaHistorial.objects.select_related('categoria').order_by(
            'categoria__orden', 'orden'
        )

class PreguntaHistorialCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = models.PreguntaHistorial
    form_class = forms.PreguntaHistorialMejoradaForm
    template_name = 'core/cuestionario/admin/pregunta_form.html'
    success_url = reverse_lazy('core:pregunta_historial_list')
    success_message = "Pregunta creada exitosamente."

class PreguntaHistorialUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = models.PreguntaHistorial
    form_class = forms.PreguntaHistorialMejoradaForm
    template_name = 'core/cuestionario/admin/pregunta_form.html'
    success_url = reverse_lazy('core:pregunta_historial_list')
    success_message = "Pregunta actualizada exitosamente."

# --- API PARA CREAR PACIENTES ---
@tenant_login_required
@require_POST
def crear_paciente_ajax(request):
    """
    API para crear pacientes vía AJAX desde el modal de la agenda
    """
    try:
        data = json.loads(request.body)
        
        # Validar datos mínimos
        required_fields = ['nombre', 'apellido', 'fecha_nacimiento']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({
                    'success': False, 
                    'error': f'El campo {field} es requerido'
                }, status=400)
        
        # Crear paciente
        paciente = models.Paciente.objects.create(
            nombre=data['nombre'],
            apellido=data['apellido'],
            email=data.get('email', ''),
            telefono=data.get('telefono', ''),
            fecha_nacimiento=data['fecha_nacimiento'],
            direccion=data.get('direccion', '')
        )
        
        return JsonResponse({
            'success': True,
            'paciente_id': paciente.id,
            'paciente_nombre': f"{paciente.nombre} {paciente.apellido}",
            'message': 'Paciente creado exitosamente'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos JSON inválidos'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al crear paciente: {str(e)}'
        }, status=500)

class CustomLogoutView(DjangoLogoutView):
    """
    Vista de logout personalizada que asegura la limpieza completa de la sesión
    """
    
    def dispatch(self, request, *args, **kwargs):
        logger.info(f"Iniciando proceso de logout para usuario: {request.user}")
        
        # Obtener el prefijo del tenant antes de limpiar la sesión
        tenant_prefix = getattr(request, 'tenant_prefix', '')
        logger.info(f"Tenant prefix detectado: '{tenant_prefix}'")
        
        # Limpiar explícitamente la sesión antes del logout oficial
        if hasattr(request, 'session'):
            logger.info(f"Sesión antes del flush: {dict(request.session.items()) if request.session else 'Sin sesion'}")
            request.session.flush()  # Elimina todos los datos de sesión y regenera la clave
            logger.info("Sesión eliminada con flush()")
        
        # Ejecutar logout de Django
        logout(request)
        logger.info("Logout de Django ejecutado")
        
        # Obtener la URL de redirección con el prefijo del tenant
        next_url = self.get_next_page()
        if not next_url:
            # Usar tenant_prefix si existe, sino usar URL normal
            next_url = f'{tenant_prefix}/accounts/login/' if tenant_prefix else reverse('login')
        
        logger.info(f"Redirigiendo a: {next_url}")
        
        # Crear respuesta de redirección
        response = HttpResponseRedirect(next_url)
        
        # ELIMINAR TODAS LAS COOKIES POSIBLES
        cookies_to_delete = ['sessionid', 'csrftoken', 'django_session', '_auth_user_id']
        for cookie_name in cookies_to_delete:
            response.delete_cookie(
                cookie_name,
                domain=None,  # Eliminar para todos los dominios
                path='/',     # Eliminar para todas las rutas
            )
            # También intentar eliminar para subdominios
            if hasattr(request, 'META') and 'HTTP_HOST' in request.META:
                host = request.META['HTTP_HOST'].split(':')[0]  # Remover puerto
                if '.' in host:
                    domain = '.' + '.'.join(host.split('.')[1:])  # .localhost
                    response.delete_cookie(cookie_name, domain=domain, path='/')
        
        # HEADERS AGRESIVOS PARA PREVENIR CACHÉ
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = 'Thu, 01 Jan 1970 00:00:00 GMT'
        response['Last-Modified'] = 'Thu, 01 Jan 1970 00:00:00 GMT'
        response['ETag'] = '""'
        response['Vary'] = '*'
        
        # Forzar cierre de conexión
        response['Connection'] = 'close'
        
        logger.info("Logout completado exitosamente")
        return response
    
    def get_next_page(self):
        """
        Obtiene la página de redirección después del logout
        """
        next_page = self.next_page
        if next_page:
            return next_page
        
        # Intentar obtener desde parámetros GET
        next_url = self.request.GET.get('next')
        if next_url:
            return next_url
            
        # Por defecto, redirigir al login con prefijo del tenant
        tenant_prefix = getattr(self.request, 'tenant_prefix', '')
        if tenant_prefix:
            return f'{tenant_prefix}/accounts/login/'
        return reverse('login')

# === FUNCIÓN DE REDIRECCIÓN RAÍZ ===
def redirect_to_dashboard(request):
    """
    Función para manejar la redirección de la raíz '/'
    Verifica autenticación y redirige apropiadamente
    Preserva parámetro tenant si existe
    """
    if not request.user.is_authenticated:
        from django.conf import settings
        login_url = settings.LOGIN_URL
        # Preservar parámetro tenant si existe
        tenant_param = request.GET.get('tenant')
        if tenant_param:
            login_url += f'?tenant={tenant_param}'
        return redirect(login_url)
    
    # Usuario autenticado, redirigir al dashboard
    from django.urls import reverse
    dashboard_url = reverse('core:dashboard')
    
    # Preservar parámetro tenant si existe
    tenant_param = request.GET.get('tenant')
    if tenant_param:
        dashboard_url += f'?tenant={tenant_param}'
    
    return redirect(dashboard_url)

# --- API PARA CANCELAR CITAS ---
@tenant_login_required
@require_POST
def cancelar_cita_api(request, cita_id):
    """
    API para cancelar una cita
    """
    try:
        cita = get_object_or_404(models.Cita, id=cita_id)
        
        # Solo permitir cancelar citas en ciertos estados
        if cita.estado not in ['PRO', 'CON']:
            return JsonResponse({
                'success': False,
                'error': 'Solo se pueden cancelar citas Programadas o Confirmadas'
            }, status=400)
        
        cita.estado = 'CAN'
        cita.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Cita cancelada exitosamente'
        }, status=200)
        
    except models.Cita.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cita no encontrada'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al cancelar cita: {str(e)}'
        }, status=500)

# === VISTAS PARA CONSENTIMIENTO INFORMADO ===

class ConsentimientoInformadoListView(TenantLoginRequiredMixin, ListView):
    """Vista para listar todos los consentimientos informados"""
    model = models.ConsentimientoInformado
    template_name = 'core/consentimiento/lista.html'
    context_object_name = 'consentimientos'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = models.ConsentimientoInformado.objects.select_related(
            'creado_por', 'actualizado_por'
        ).order_by('-creado_en')
        
        # Filtros
        tipo = self.request.GET.get('tipo')
        estado = self.request.GET.get('estado')
        vigente = self.request.GET.get('vigente')
        
        if tipo:
            queryset = queryset.filter(tipo_documento=tipo)
            
        if estado:
            queryset = queryset.filter(estado=estado)
            
        if vigente == '1':
            from datetime import date
            today = date.today()
            queryset = queryset.filter(
                estado='ACTIVO',
                fecha_vigencia_inicio__lte=today
            ).filter(
                models.Q(fecha_vigencia_fin__isnull=True) |
                models.Q(fecha_vigencia_fin__gte=today)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Opciones para filtros
        context['tipos_documento'] = models.ConsentimientoInformado.TIPOS_DOCUMENTO
        context['estados'] = models.ConsentimientoInformado.ESTADOS
        
        # Filtros activos
        context['filtros_activos'] = {
            'tipo': self.request.GET.get('tipo'),
            'estado': self.request.GET.get('estado'),
            'vigente': self.request.GET.get('vigente'),
        }
        
        return context

class ConsentimientoInformadoDetailView(TenantLoginRequiredMixin, DetailView):
    """Vista para mostrar detalles de un consentimiento informado"""
    model = models.ConsentimientoInformado
    template_name = 'core/consentimiento/detalle.html'
    context_object_name = 'consentimiento'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        consentimiento = self.object
        
        # Obtener firmas asociadas
        context['pacientes_firmantes'] = models.PacienteConsentimiento.objects.filter(
            consentimiento=consentimiento
        ).select_related('paciente', 'presentado_por').order_by('-fecha_presentado')
        
        # Estadísticas de uso
        total_presentados = context['pacientes_firmantes'].count()
        firmados = context['pacientes_firmantes'].filter(estado='FIRMADO').count()
        pendientes = context['pacientes_firmantes'].filter(estado='PENDIENTE').count()
        rechazados = context['pacientes_firmantes'].filter(estado='RECHAZADO').count()
        
        context['estadisticas'] = {
            'total_presentados': total_presentados,
            'firmados': firmados,
            'pendientes': pendientes,
            'rechazados': rechazados,
            'porcentaje_firmados': (firmados / total_presentados * 100) if total_presentados > 0 else 0
        }
        
        return context

class ConsentimientoInformadoCreateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, CreateView):
    """Vista para crear un nuevo consentimiento informado"""
    model = models.ConsentimientoInformado
    template_name = 'core/consentimiento/form.html'
    fields = [
        'titulo', 'tipo_documento', 'descripcion', 'archivo_pdf',
        'version', 'fecha_vigencia_inicio', 'fecha_vigencia_fin',
        'cumple_cofepris', 'requiere_testigos', 'estado'
    ]
    success_url = reverse_lazy('core:consentimiento_list')
    success_message = "Consentimiento informado '%(titulo)s' creado exitosamente."
    
    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        return super().form_valid(form)

class ConsentimientoInformadoUpdateView(TenantSuccessUrlMixin, TenantLoginRequiredMixin, SuccessMessageMixin, UpdateView):
    """Vista para actualizar un consentimiento informado"""
    model = models.ConsentimientoInformado
    template_name = 'core/consentimiento/form.html'
    fields = [
        'titulo', 'tipo_documento', 'descripcion', 'archivo_pdf',
        'version', 'fecha_vigencia_inicio', 'fecha_vigencia_fin',
        'cumple_cofepris', 'requiere_testigos', 'estado'
    ]
    success_url = reverse_lazy('core:consentimiento_list')
    success_message = "Consentimiento informado '%(titulo)s' actualizado exitosamente."
    
    def form_valid(self, form):
        form.instance.actualizado_por = self.request.user
        return super().form_valid(form)

class PacienteConsentimientoListView(TenantLoginRequiredMixin, ListView):
    """Vista para listar consentimientos de pacientes"""
    model = models.PacienteConsentimiento
    template_name = 'core/consentimiento/paciente_lista.html'
    context_object_name = 'consentimientos_pacientes'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = models.PacienteConsentimiento.objects.select_related(
            'paciente', 'consentimiento', 'presentado_por', 'cita'
        ).order_by('-fecha_presentado')
        
        # Filtros
        estado = self.request.GET.get('estado')
        paciente_busqueda = self.request.GET.get('paciente')
        tipo = self.request.GET.get('tipo')
        
        if estado:
            queryset = queryset.filter(estado=estado)
            
        if paciente_busqueda:
            from django.db.models import Q
            # Buscar por nombre, apellido o email del paciente
            queryset = queryset.filter(
                Q(paciente__nombre__icontains=paciente_busqueda) |
                Q(paciente__apellido__icontains=paciente_busqueda) |
                Q(paciente__email__icontains=paciente_busqueda)
            )
            
        if tipo:
            queryset = queryset.filter(consentimiento__tipo_documento=tipo)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Opciones para filtros
        context['estados_firma'] = models.PacienteConsentimiento.ESTADOS_FIRMA
        context['tipos_documento'] = models.ConsentimientoInformado.TIPOS_DOCUMENTO
        
        # Filtros activos
        context['filtros_activos'] = {
            'estado': self.request.GET.get('estado'),
            'paciente': self.request.GET.get('paciente'),
            'tipo': self.request.GET.get('tipo'),
        }
        
        return context

class PacienteConsentimientoDetailView(TenantLoginRequiredMixin, DetailView):
    """Vista para mostrar detalles de un consentimiento de paciente"""
    model = models.PacienteConsentimiento
    template_name = 'core/consentimiento/paciente_detalle.html'
    context_object_name = 'paciente_consentimiento'

@tenant_login_required
def presentar_consentimiento_desde_cuestionario(request, cuestionario_id):
    """Vista para presentar consentimiento informado desde un cuestionario completado"""
    cuestionario = get_object_or_404(models.CuestionarioCompletado, id=cuestionario_id)
    
    # Verificar si ya se presentó un consentimiento
    if cuestionario.consentimiento_presentado:
        messages.info(
            request, 
            'Ya se ha presentado un consentimiento informado para este cuestionario.'
        )
        return redirect('core:paciente_consentimiento_detail', pk=cuestionario.consentimiento_presentado.id)
    
    if request.method == 'POST':
        tipo_documento = request.POST.get('tipo_documento')
        dentista_id = request.POST.get('dentista')
        
        # Obtener dentista si se especifica
        dentista = None
        if dentista_id:
            try:
                dentista = models.PerfilDentista.objects.get(id=dentista_id)
            except models.PerfilDentista.DoesNotExist:
                messages.error(request, 'Dentista no encontrado.')
                return redirect('core:cuestionario_detalle', paciente_id=cuestionario.paciente.id)
        
        # Presentar consentimiento
        paciente_consentimiento = cuestionario.presentar_consentimiento(
            tipo_documento=tipo_documento,
            dentista=dentista
        )
        
        if paciente_consentimiento:
            messages.success(
                request,
                f'Consentimiento informado presentado al paciente {cuestionario.paciente}.'
            )
            return redirect('core:paciente_consentimiento_detail', pk=paciente_consentimiento.id)
        else:
            messages.error(
                request,
                'No se pudo presentar el consentimiento. Verifique que existan documentos vigentes.'
            )
    
    # GET: Mostrar formulario
    context = {
        'cuestionario': cuestionario,
        'tipos_documento': models.ConsentimientoInformado.TIPOS_DOCUMENTO,
        'dentistas': models.PerfilDentista.objects.filter(activo=True).order_by('apellido', 'nombre'),
        'tipo_recomendado': cuestionario.determinar_tipo_consentimiento(),
    }
    
    return render(request, 'core/consentimiento/presentar_desde_cuestionario.html', context)

@tenant_login_required
def firmar_consentimiento(request, paciente_consentimiento_id):
    """Vista para que un paciente firme un consentimiento informado"""
    paciente_consentimiento = get_object_or_404(
        models.PacienteConsentimiento, 
        id=paciente_consentimiento_id
    )
    
    if paciente_consentimiento.estado != 'PENDIENTE':
        messages.warning(
            request, 
            'Este consentimiento ya ha sido procesado y no puede ser modificado.'
        )
        return redirect('core:paciente_consentimiento_detail', pk=paciente_consentimiento_id)
    
    if request.method == 'POST':
        # Procesar firmas (esto se implementaría con JavaScript para capturar firmas)
        firma_paciente = request.FILES.get('firma_paciente')
        
        if firma_paciente:
            paciente_consentimiento.firma_paciente = firma_paciente
            
            # Si requiere testigos, procesar firmas de testigos
            if paciente_consentimiento.consentimiento.requiere_testigos:
                nombre_testigo1 = request.POST.get('nombre_testigo1')
                firma_testigo1 = request.FILES.get('firma_testigo1')
                
                if nombre_testigo1 and firma_testigo1:
                    paciente_consentimiento.nombre_testigo1 = nombre_testigo1
                    paciente_consentimiento.firma_testigo1 = firma_testigo1
                else:
                    messages.error(request, 'Se requiere nombre y firma del testigo.')
                    return redirect('core:firmar_consentimiento', paciente_consentimiento_id=paciente_consentimiento_id)
            
            # Marcar como firmado
            if paciente_consentimiento.marcar_como_firmado():
                messages.success(request, 'Consentimiento firmado exitosamente.')
                return redirect('core:paciente_consentimiento_detail', pk=paciente_consentimiento_id)
        else:
            messages.error(request, 'Se requiere la firma del paciente.')
    
    context = {
        'paciente_consentimiento': paciente_consentimiento,
    }
    
    return render(request, 'core/consentimiento/firmar.html', context)

@tenant_login_required
def descargar_consentimiento_pdf(request, consentimiento_id):
    """Vista para descargar el PDF de un consentimiento informado"""
    consentimiento = get_object_or_404(models.ConsentimientoInformado, id=consentimiento_id)
    
    try:
        response = FileResponse(
            consentimiento.archivo_pdf.open('rb'),
            content_type='application/pdf'
        )
        filename = f"consentimiento_{consentimiento.tipo_documento}_{consentimiento.version}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except FileNotFoundError:
        messages.error(request, 'El archivo PDF no se encuentra disponible.')
        return redirect('core:consentimiento_detail', pk=consentimiento_id)

# === INTEGRACIÓN CON CUESTIONARIO ===

@tenant_login_required
def cuestionario_con_consentimiento(request, paciente_id):
    """Vista integrada que muestra el cuestionario y permite acceder al consentimiento"""
    paciente = get_object_or_404(models.Paciente, id=paciente_id)
    
    # Obtener último cuestionario
    ultimo_cuestionario = paciente.cuestionarios_completados.first()
    
    context = {
        'paciente': paciente,
        'ultimo_cuestionario': ultimo_cuestionario,
    }
    
    # Si hay cuestionario, agregar información de consentimiento
    if ultimo_cuestionario:
        context['requiere_consentimiento'] = ultimo_cuestionario.requiere_consentimiento_informado()
        context['estado_consentimiento'] = ultimo_cuestionario.estado_consentimiento()
        context['consentimiento_presentado'] = ultimo_cuestionario.consentimiento_presentado
        context['puede_proceder'], context['mensaje_tratamiento'] = ultimo_cuestionario.puede_proceder_con_tratamiento()
    
    return render(request, 'core/cuestionario/cuestionario_con_consentimiento.html', context)

@tenant_login_required
def prueba_boca_abierta(request):
    """Vista de prueba para el odontograma con forma de boca abierta ovalada"""
    return render(request, 'prueba_boca_abierta.html')

@tenant_login_required
def prueba_fdi_universal(request):
    """Vista de prueba para el odontograma con Sistema de Numeración FDI Universal"""
    return render(request, 'prueba_fdi_universal.html')
